"""
core/helpers.py — Lógica de negocio compartida entre routers.

Incluye:
- hash_password, init_db
- crear_notificacion, _registrar_audit
- get_visibility_filter, can_edit_registro
- helpers de prestadores y usuarios (carga masiva Excel)
- _reports (almacén temporal de Excel descargables)
"""
import hashlib
import io
import json
import sqlite3
import time
import uuid
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd

from app.config import DB_PATH, DATABASE_URL, ADMIN_INITIAL_PASSWORD, DEFAULT_SUPERIOR_INMEDIATO
from app.core.seed_data import (
    CAMPOS_DEFAULT, LISTAS_DEFAULT,
    CIUDAD_CODIGOS_DEFAULT, FESTIVOS_DEFAULT,
)
from app.core.fields import (
    _col, _campo_sql_type, _migrate_registros, _refresh_globals,
    CODE_TO_COLNAME, ALL_FIELD_CODES, ALL_FIELD_CODES_SET,
    ROLES_FIELDS, NOMBRE_TO_CODE,
)


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
import bcrypt as _bcrypt


def hash_password(password: str) -> str:
    return _bcrypt.hashpw(password.encode("utf-8"), _bcrypt.gensalt(rounds=12)).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    if hashed.startswith("$2b$") or hashed.startswith("$2a$"):
        return _bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    # Fallback SHA-256 para hashes legacy — se re-hashean con bcrypt en el próximo login
    return hashlib.sha256(plain.encode()).hexdigest() == hashed


# ---------------------------------------------------------------------------
# Almacén temporal de reportes (token → (timestamp, bytes))
# ---------------------------------------------------------------------------
_reports: dict = {}


def _store_report(excel_bytes: bytes, db=None) -> str:
    """Guarda el reporte en la tabla temp_reports (DB) o en memoria como fallback."""
    token = str(uuid.uuid4())
    if db is not None:
        try:
            db.execute("DELETE FROM temp_reports WHERE created_at < ?", (time.time() - 600,))
            db.execute(
                "INSERT INTO temp_reports (token, created_at, data) VALUES (?, ?, ?)",
                (token, time.time(), excel_bytes),
            )
            db.commit()
            return token
        except Exception:
            pass  # fallback a memoria si la tabla aún no existe
    # Fallback: almacén en memoria
    now = time.time()
    expired = [k for k, (ts, _) in list(_reports.items()) if now - ts > 600]
    for k in expired:
        del _reports[k]
    _reports[token] = (now, excel_bytes)
    return token


# ---------------------------------------------------------------------------
# Notificaciones
# ---------------------------------------------------------------------------
def crear_notificacion(db, usuario_destino: str, tipo: str, mensaje: str,
                       registro_id=None):
    """Inserta una notificación. No hace commit."""
    db.execute(
        "INSERT INTO notificaciones (usuario_destino, tipo, mensaje, registro_id, fecha_creacion)"
        " VALUES (?, ?, ?, ?, ?)",
        (usuario_destino, tipo, mensaje, registro_id, datetime.now().isoformat()),
    )


# ---------------------------------------------------------------------------
# Auditoría
# ---------------------------------------------------------------------------
def _registrar_audit(db, registro_id, accion: str, old_row, valid_campos: dict,
                     es_autorizado: bool, sess: dict, motivo: str = ""):
    """Registra en audit_log.
    - Autorizados: reemplaza fila existente (UPSERT).
    - Terceros y eliminaciones: siempre INSERT.
    Recibe `sess` (dict de sesión) en lugar de usar session global de Flask.
    motivo: comentario de cierre (motivo_comentario).
    """
    usuario = sess.get("usuario", "?")
    rol     = ", ".join(sess.get("permisos", [sess.get("rol", "?")]))

    diff: dict = {}
    if accion == "eliminacion":
        for cod, col in CODE_TO_COLNAME.items():
            try:
                val = old_row[col]
                if val is not None and val != "":
                    diff[col] = {"antes": str(val), "despues": None}
            except (IndexError, KeyError):
                pass
    else:
        for cod, new_val in valid_campos.items():
            col = _col(cod)
            try:
                old_val = old_row[col]
            except (IndexError, KeyError):
                old_val = None
            if str(old_val or "") != str(new_val or ""):
                diff[col] = {
                    "antes":   str(old_val)  if old_val  is not None else "",
                    "despues": str(new_val)  if new_val  is not None else "",
                }

    if not diff and accion != "eliminacion":
        return  # sin cambios reales

    consecutivo = ""
    nombre_prestador = ""
    try:
        consecutivo      = old_row[_col("A")] or ""
        nombre_prestador = old_row[_col("I")] or ""
    except Exception:
        pass

    fecha     = datetime.now().isoformat()
    diff_json = json.dumps(diff, ensure_ascii=False)

    if es_autorizado and accion != "eliminacion":
        db.execute(
            "DELETE FROM audit_log WHERE registro_id = ? AND es_autorizado = 1",
            (registro_id,)
        )

    db.execute(
        "INSERT INTO audit_log "
        "(registro_id, accion, usuario, rol, es_autorizado, campos_diff, "
        "consecutivo, nombre_prestador, fecha, motivo_comentario)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (registro_id, accion, usuario, rol, 1 if es_autorizado else 0,
         diff_json, consecutivo, nombre_prestador, fecha, motivo or "")
    )


# ---------------------------------------------------------------------------
# Filtros de visibilidad y edición
# ---------------------------------------------------------------------------
def get_visibility_filter(db, sess: dict):
    """Retorna (where_clause, params) para filtrar registros visibles.
    Recibe `sess` (dict de sesión activa) en lugar de session global de Flask.

    Reglas:
    - ADMIN / CONTRALOR: ven todo.
    - LIDER: ve registros creados por él, asignados a él (AG) y los de sus subordinados directos.
    - GESTOR 1 / GESTOR 2 / demás: ve registros creados por él y los asignados a él (AG).
    """
    if sess.get("is_admin"):
        return None, None
    if "CONTRALOR" in sess.get("permisos", []):
        return None, None

    usuario  = sess["usuario"]
    permisos = sess.get("permisos", [])
    ag_col   = _col("AG")

    me_row = db.execute(
        "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
        (usuario,)
    ).fetchone()
    nombre_completo = (me_row["nombre_completo"] or "").strip() if me_row else None

    conditions: list = []
    params:     list = []

    # Registros creados por mí
    conditions.append("usuario = ?")
    params.append(usuario)

    # Registros asignados a mí (AG = nombre_completo, insensible a mayúsculas/espacios)
    if nombre_completo:
        conditions.append(f"LOWER(TRIM({ag_col})) = LOWER(?)")
        params.append(nombre_completo)

    # LÍDER: también ve los registros creados por sus subordinados directos
    # o asignados a ellos en AG (cubre registros históricos con usuario='importacion')
    if "LIDER" in permisos:
        sub_rows = db.execute(
            "SELECT usuario, nombre_completo FROM usuarios WHERE superior_inmediato = ? AND activo = 1",
            (usuario,)
        ).fetchall()
        sub_users  = [r["usuario"] for r in sub_rows]
        sub_nombres = [r["nombre_completo"].strip() for r in sub_rows if r["nombre_completo"]]
        if sub_users:
            ph = ",".join("?" * len(sub_users))
            conditions.append(f"usuario IN ({ph})")
            params.extend(sub_users)
        if sub_nombres:
            ph = ",".join("?" * len(sub_nombres))
            conditions.append(f"LOWER(TRIM({ag_col})) IN ({ph})")
            params.extend(n.lower() for n in sub_nombres)

    # Siempre ve los registros donde tiene una auditoría activa asignada como destinatario,
    # aunque el registro haya migrado de sección (cerrado, by_env, etc.).
    conditions.append(
        "id IN (SELECT registro_id FROM auditoria_registros"
        " WHERE destinatario_usuario = ? AND estado IN ('activa','en_proceso'))"
    )
    params.append(usuario)

    where = "(" + " OR ".join(conditions) + ")"
    return where, tuple(params)


def _ciudad_del_registro(row) -> Optional[str]:
    try:
        return row[_col("C")] or None
    except (IndexError, KeyError):
        return None


def can_edit_registro(row, sess: dict, db=None) -> bool:
    """Determina si el usuario en sesión puede editar el registro.

    Reglas:
    - ADMIN / CONTRALOR: pueden editar todo, salvo estado_aprobacion_n='pendiente'/'cancelado'.
    - GESTOR 1 / GESTOR 2: pueden editar registros que crearon o que tienen asignados (AG).
      El frontend restringe a la sección propia si crearon pero no están asignados.
    - LÍDER: puede editar registros que creó, que tiene asignados (AG) o
      que fueron creados por sus subordinados directos.
    """
    # Bloqueo por estado_aprobacion_n (aplica antes de cualquier rol):
    # - 'pendiente': nadie puede editar (solo endpoints dedicados)
    # - 'cancelado': nadie puede editar
    # - 'rechazado': solo CONTRALOR/ADMIN pueden editar
    try:
        _estado_n = (row["estado_aprobacion_n"] or "").strip()
    except (KeyError, IndexError):
        _estado_n = ""

    if _estado_n in ("pendiente", "cancelado"):
        return False

    if _estado_n == "rechazado":
        if not sess.get("is_admin") and "CONTRALOR" not in sess.get("permisos", []):
            return False

    if sess.get("is_admin"):
        return True
    if "CONTRALOR" in sess.get("permisos", []):
        return True

    # Registro cerrado: solo LÍDER puede editar (CONTRALOR/ADMIN ya retornaron True arriba)
    _AC_CIERRE_VALS_SET = {
        "CERRADO POR CANCELACION DE MESA",
        "CERRADO POR CANCELACION DE MESAS",
        "CERRADO SIN FINALIZACIÓN",
        "IPS NO ASISTE A MESAS",
    }
    _ac_col = _col("AC")
    _bd_col = _col("BD")
    _ce_col = _col("CE")
    try:
        _ac_val = (row[_ac_col] or "").strip()
    except (IndexError, KeyError):
        _ac_val = ""
    try:
        _bd_val = (row[_bd_col] or "").strip()
    except (IndexError, KeyError):
        _bd_val = ""
    try:
        _ce_val = (row[_ce_col] or "").strip()
    except (IndexError, KeyError):
        _ce_val = ""
    _CE_CIERRE_VALS = {"CERRADO POR CANCELACION DE MESAS", "CERRADO SIN FINALIZACIÓN"}
    _registro_cerrado = (
        _ac_val in _AC_CIERRE_VALS_SET or
        _bd_val == "CERRADO SIN FINALIZACION" or
        _ce_val in _CE_CIERRE_VALS
    )
    if _registro_cerrado and "LIDER" not in sess.get("permisos", []):
        return False

    usuario  = sess["usuario"]
    permisos = sess.get("permisos", [])

    # El creador siempre puede editar
    # (el frontend limita a su propia sección si no está asignado al registro)
    if row["usuario"] == usuario:
        return True

    if db is not None:
        ag_col = _col("AG")
        try:
            ag_val = (row[ag_col] or "").strip()
        except (IndexError, KeyError):
            ag_val = ""

        me_row = db.execute(
            "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
            (usuario,)
        ).fetchone()
        nombre_completo = (me_row["nombre_completo"] or "").strip() if me_row else None

        # Asignado a mí (AG = nombre_completo, insensible a mayúsculas/espacios)
        if nombre_completo and ag_val.lower() == nombre_completo.lower():
            return True

        # LÍDER: puede editar registros de sus subordinados directos
        if "LIDER" in permisos:
            sub_row = db.execute(
                "SELECT 1 FROM usuarios WHERE superior_inmediato = ? AND usuario = ? AND activo = 1",
                (usuario, row["usuario"])
            ).fetchone()
            if sub_row:
                return True
            # También puede editar registros asignados a un subordinado directo (campo AG),
            # aunque el creador sea 'importacion' (datos históricos) u otro usuario
            # externo a la jerarquía. Simétrico con get_visibility_filter que ya muestra
            # estos registros al LIDER.
            if ag_val:
                sub_ag = db.execute(
                    "SELECT 1 FROM usuarios"
                    " WHERE superior_inmediato = ? AND activo = 1"
                    " AND LOWER(TRIM(nombre_completo)) = LOWER(?)",
                    (usuario, ag_val),
                ).fetchone()
                if sub_ag:
                    return True

    return False


# ---------------------------------------------------------------------------
# Helpers de Excel — Prestadores
# ---------------------------------------------------------------------------
_PREST_COL_NAMES = [
    'NUM_ID', 'CODIGO_COMPANIA', 'COMPANIA', 'COD_PLAN', 'DESCRIPCION_PLAN',
    'FORMA_CONTRATACION', 'DIGITO_VERIFICACION', 'TIPO_ID', 'TIPO_PERSONA',
    'RELACION_EPS', 'NOMBRE_SUCURSAL', 'CODIGO_SUCURSAL', 'CIUDAD_COD_DANE',
    'DESCRIPCION_CIUDAD', 'DEPARTAMENTO', 'REGIONAL', 'ESPECIALIDAD',
    'DESCRIPCION_ESPECIALIDAD', 'ESTADO', 'TIPO_CONVENIO', 'DIRECCION',
    'TELEFONO_1', 'EXTENSION_1', 'TELEFONO_2', 'EXTENSION_2', 'CORREO',
    'FECHA_INICIO_PORTABILIDAD', 'FECHA_FIN_PORTABILIDAD', 'COD_HABILITACION',
    'HABILITACION_SEDE', 'FECHA_INICIO_HABILITACION', 'FECHA_VENCIMIENTO_HABILITACION',
    'NUMERO_CONTRATO', 'FECHA_INICIO_CONVENIO', 'FECHA_FIN_CONVENIO',
    'TIPO_PRESTADOR', 'NATURALEZA_IPS', 'TIPO_ATENCION', 'PREMIUM',
    'GLOSA_SOSTENIDA', 'PRIORIDAD_SERVICIO',
]

_USER_COL_NAMES = [
    'USUARIO', 'NOMBRE', 'CEDULA', 'CARGO', 'CORREO ELECTRONICO',
    'REGIONAL', 'GESTOR_1', 'GESTOR_2', 'LIDER', 'CONTRALOR',
    'SUPERIOR_INMEDIATO', 'ACTIVO', 'CONTRASENA',
]
_USER_REPORT_COL_NAMES = [c for c in _USER_COL_NAMES if c != 'CONTRASENA']


def _sv(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    return s if s else None


def _dv_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        s = str(val).strip()
        return s if s else None


def _nit_norm(raw):
    if raw is None:
        return None
    try:
        return str(int(float(raw))).strip()
    except (ValueError, TypeError):
        s = str(raw).strip()
        return s if s else None


def _int_perm(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0
    try:
        return min(1, max(0, int(float(val))))
    except (ValueError, TypeError):
        return 0


def _row_fields(row):
    return (
        _sv(row['CODIGO_COMPANIA']), _sv(row['COMPANIA']),
        _sv(row['COD_PLAN']), _sv(row['DESCRIPCION_PLAN']),
        _sv(row['FORMA_CONTRATACION']), _dv_str(row['DIGITO_VERIFICACION']),
        _sv(row['TIPO_ID']), _sv(row['TIPO_PERSONA']),
        _sv(row['RELACION_EPS']), _sv(row['NOMBRE_SUCURSAL']),
        _sv(row['CODIGO_SUCURSAL']), _sv(row['CIUDAD_COD_DANE']),
        _sv(row['DESCRIPCION_CIUDAD']), _sv(row['DEPARTAMENTO']),
        _sv(row['REGIONAL']), _sv(row['ESPECIALIDAD']),
        _sv(row['DESCRIPCION_ESPECIALIDAD']), _sv(row['ESTADO']),
        _sv(row['TIPO_CONVENIO']), _sv(row['DIRECCION']),
        _sv(row['TELEFONO_1']), _sv(row['EXTENSION_1']),
        _sv(row['TELEFONO_2']), _sv(row['EXTENSION_2']),
        _sv(row['CORREO']), _sv(row['FECHA_INICIO_PORTABILIDAD']),
        _sv(row['FECHA_FIN_PORTABILIDAD']), _sv(row['COD_HABILITACION']),
        _sv(row['HABILITACION_SEDE']), _sv(row['FECHA_INICIO_HABILITACION']),
        _sv(row['FECHA_VENCIMIENTO_HABILITACION']), _sv(row['NUMERO_CONTRATO']),
        _sv(row['FECHA_INICIO_CONVENIO']), _sv(row['FECHA_FIN_CONVENIO']),
        _sv(row['TIPO_PRESTADOR']), _sv(row['NATURALEZA_IPS']),
        _sv(row['TIPO_ATENCION']), _sv(row['PREMIUM']),
        _sv(row['GLOSA_SOSTENIDA']), _sv(row['PRIORIDAD_SERVICIO']),
    )


def _build_report_excel(rows, motivos, col_names=None) -> io.BytesIO:
    if col_names is None:
        col_names = _PREST_COL_NAMES
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros rechazados"
    headers = col_names + ["Motivo"]
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for row_idx, (row_data, motivo) in enumerate(zip(rows, motivos), 2):
        for col_idx, col_name in enumerate(col_names, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))
        ws.cell(row=row_idx, column=len(col_names) + 1, value=motivo)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _existing_nits_map(db) -> dict:
    mapping = {}
    for row in db.execute("SELECT DISTINCT nit FROM prestadores").fetchall():
        # Usar índice 0 para compatibilidad con sqlite3.Row y _PgRow (dict).
        # El desempaquetado "for (raw,) in rows" falla en PostgreSQL porque
        # _PgRow es un dict y la iteración devuelve las claves, no los valores.
        raw = row[0]
        norm = _nit_norm(raw)
        if norm:
            mapping[norm] = raw
    return mapping


def _validate_and_parse_prestadores(file_obj):
    try:
        df = pd.read_excel(file_obj)
        df = df.where(pd.notna(df), None)
    except Exception as e:
        return None, f"No se pudo leer el archivo Excel: {e}"
    if df.shape[1] != len(_PREST_COL_NAMES):
        return None, (
            f"El archivo tiene {df.shape[1]} columnas pero se esperan "
            f"{len(_PREST_COL_NAMES)}. Descarga el formato correcto."
        )
    df.columns = _PREST_COL_NAMES
    return df, None


def _validate_and_parse_usuarios(file_obj):
    try:
        df = pd.read_excel(file_obj)
        df = df.where(pd.notna(df), None)
    except Exception as e:
        return None, f"No se pudo leer el archivo Excel: {e}"
    if df.shape[1] != len(_USER_COL_NAMES):
        return None, (
            f"El archivo tiene {df.shape[1]} columnas pero se esperan "
            f"{len(_USER_COL_NAMES)}. Descarga el formato correcto."
        )
    df.columns = _USER_COL_NAMES
    return df, None


def _admin_count(db) -> int:
    return db.execute(
        "SELECT COUNT(*) FROM usuarios WHERE is_admin=1 AND activo=1"
    ).fetchone()[0]


# SQL reutilizable para prestadores
_INSERT_PREST_SQL = """
    INSERT INTO prestadores
       (nit, codigo_compania, compania, cod_plan, descripcion_plan,
        forma_contratacion, digito_verificacion, tipo_id, tipo_persona,
        relacion_eps, nombre_sucursal, codigo_sucursal, ciudad_cod_dane,
        ciudad, departamento, regional, especialidad, descripcion_especialidad,
        estado, tipo_convenio, direccion, telefono, extension_1, telefono_2,
        extension_2, correo, fecha_inicio_portabilidad, fecha_fin_portabilidad,
        cod_habilitacion, habilitacion_sede, fecha_inicio_habilitacion,
        fecha_vencimiento_habilitacion, numero_contrato, fecha_inicio_convenio,
        fecha_fin_convenio, tipo_prestador, naturaleza_ips, tipo_atencion,
        premium, glosa_sostenida, prioridad_servicio, creado_manual)
    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,0)
"""

_UPDATE_PREST_SQL = """
    UPDATE prestadores SET
        nit=?, codigo_compania=?, compania=?, cod_plan=?, descripcion_plan=?,
        forma_contratacion=?, digito_verificacion=?, tipo_id=?, tipo_persona=?,
        relacion_eps=?, nombre_sucursal=?, codigo_sucursal=?, ciudad_cod_dane=?,
        ciudad=?, departamento=?, regional=?, especialidad=?,
        descripcion_especialidad=?, estado=?, tipo_convenio=?, direccion=?,
        telefono=?, extension_1=?, telefono_2=?, extension_2=?, correo=?,
        fecha_inicio_portabilidad=?, fecha_fin_portabilidad=?, cod_habilitacion=?,
        habilitacion_sede=?, fecha_inicio_habilitacion=?,
        fecha_vencimiento_habilitacion=?, numero_contrato=?,
        fecha_inicio_convenio=?, fecha_fin_convenio=?,
        tipo_prestador=?, naturaleza_ips=?, tipo_atencion=?,
        premium=?, glosa_sostenida=?, prioridad_servicio=?, creado_manual=0
    WHERE nit=?
"""


def sync_prestadores_from_excel(file_obj, db) -> dict:
    df, err = _validate_and_parse_prestadores(file_obj)
    if err:
        return {"error": err, "updated": 0, "inserted": 0, "errors": [], "report_token": None}

    nit_map      = _existing_nits_map(db)
    existing_nits = set(nit_map.keys())
    error_rows, error_motivos = [], []
    updated = inserted = 0

    nit_rows: dict = {}
    for _, row in df.iterrows():
        nit = _nit_norm(row['NUM_ID'])
        if not nit:
            error_rows.append({c: _sv(row.get(c)) for c in _PREST_COL_NAMES})
            error_motivos.append("NIT vacío o inválido")
        else:
            nit_rows[nit] = row

    for nit, row in nit_rows.items():
        fields = _row_fields(row)
        try:
            if nit in existing_nits:
                db.execute(_UPDATE_PREST_SQL, (nit, *fields, nit_map[nit]))
                updated += 1
            else:
                db.execute(_INSERT_PREST_SQL, (nit, *fields))
                inserted += 1
        except Exception as exc:
            error_rows.append({c: _sv(row.get(c)) for c in _PREST_COL_NAMES})
            error_motivos.append(f"Error de base de datos: {exc}")

    db.commit()
    report_token = None
    if error_rows:
        buf = _build_report_excel(error_rows, error_motivos)
        report_token = _store_report(buf.getvalue(), db)

    return {
        "updated": updated, "inserted": inserted,
        "errors": [
            {"nit": r.get("NUM_ID"), "nombre": r.get("NOMBRE_SUCURSAL"), "motivo": m}
            for r, m in zip(error_rows, error_motivos)
        ],
        "report_token": report_token,
    }


def cargar_prestadores(file_obj, db) -> dict:
    df, err = _validate_and_parse_prestadores(file_obj)
    if err:
        return {"error": err, "inserted": 0, "duplicates": [], "errors": [], "report_token": None}

    nit_map       = _existing_nits_map(db)
    existing_nits = set(nit_map.keys())
    dup_rows, dup_motivos   = [], []
    error_rows, error_motivos = [], []
    seen_in_file: set = set()
    inserted = 0

    for _, row in df.iterrows():
        nit      = _nit_norm(row['NUM_ID'])
        row_dict = {c: _sv(row.get(c)) for c in _PREST_COL_NAMES}
        if not nit:
            error_rows.append(row_dict); error_motivos.append("NIT vacío o inválido")
            continue
        if nit in existing_nits:
            dup_rows.append(row_dict); dup_motivos.append(f"NIT {nit} ya existe en la BD")
            continue
        if nit in seen_in_file:
            dup_rows.append(row_dict); dup_motivos.append(f"NIT {nit} duplicado en el archivo")
            continue
        seen_in_file.add(nit)
        try:
            db.execute(_INSERT_PREST_SQL, (nit, *_row_fields(row)))
            inserted += 1
        except Exception as exc:
            error_rows.append(row_dict); error_motivos.append(f"Error de BD: {exc}")

    db.commit()
    rejected = dup_rows + error_rows
    report_token = None
    if rejected:
        buf = _build_report_excel(rejected, dup_motivos + error_motivos)
        report_token = _store_report(buf.getvalue(), db)

    return {
        "inserted": inserted,
        "duplicates": [
            {"nit": r.get("NUM_ID"), "nombre": r.get("NOMBRE_SUCURSAL"), "motivo": m}
            for r, m in zip(dup_rows, dup_motivos)
        ],
        "errors": [
            {"nit": r.get("NUM_ID"), "nombre": r.get("NOMBRE_SUCURSAL"), "motivo": m}
            for r, m in zip(error_rows, error_motivos)
        ],
        "report_token": report_token,
    }


def cargar_usuarios(file_obj, db) -> dict:
    df, err = _validate_and_parse_usuarios(file_obj)
    if err:
        return {"error": err, "inserted": 0, "duplicates": [], "errors": [], "report_token": None}

    existing = {r[0]: r[1] for r in db.execute(
        "SELECT usuario, is_admin FROM usuarios"
    ).fetchall()}

    dup_rows, dup_motivos   = [], []
    error_rows, error_motivos = [], []
    seen_in_file: set = set()
    inserted = 0

    for _, row in df.iterrows():
        usuario   = _sv(row['USUARIO'])
        contrasena = _sv(row['CONTRASENA'])
        row_rep   = {c: _sv(row.get(c)) for c in _USER_REPORT_COL_NAMES}

        if not usuario:
            error_rows.append(row_rep); error_motivos.append("Campo USUARIO vacío o inválido"); continue
        if not contrasena:
            error_rows.append(row_rep); error_motivos.append(f"CONTRASENA vacía para '{usuario}'"); continue
        if usuario in existing:
            if existing[usuario] == 1:
                error_rows.append(row_rep)
                error_motivos.append(f"'{usuario}' es Administrador y no puede modificarse por carga masiva")
            else:
                dup_rows.append(row_rep); dup_motivos.append(f"Usuario '{usuario}' ya existe en la BD")
            continue
        if usuario in seen_in_file:
            dup_rows.append(row_rep); dup_motivos.append(f"Usuario '{usuario}' duplicado en el archivo"); continue
        seen_in_file.add(usuario)
        try:
            db.execute(
                """INSERT INTO usuarios
                   (usuario, password_hash, nombre_completo, cedula, cargo, correo, regional,
                    perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor,
                    is_admin, superior_inmediato, activo)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)""",
                (usuario, hash_password(contrasena),
                 _sv(row['NOMBRE']), _sv(row['CEDULA']),
                 _sv(row['CARGO']) or "Otro", _sv(row['CORREO ELECTRONICO']),
                 _sv(row['REGIONAL']),
                 _int_perm(row['GESTOR_1']), _int_perm(row['GESTOR_2']),
                 _int_perm(row['LIDER']), _int_perm(row['CONTRALOR']),
                 _sv(row['SUPERIOR_INMEDIATO']) or DEFAULT_SUPERIOR_INMEDIATO or None,
                 _int_perm(row['ACTIVO'])),
            )
            inserted += 1
        except Exception as exc:
            error_rows.append(row_rep); error_motivos.append(f"Error de BD: {exc}")

    db.commit()
    rejected = dup_rows + error_rows
    report_token = None
    if rejected:
        buf = _build_report_excel(rejected, dup_motivos + error_motivos, _USER_REPORT_COL_NAMES)
        report_token = _store_report(buf.getvalue(), db)

    return {
        "inserted": inserted,
        "duplicates": [
            {"usuario": r.get("USUARIO"), "nombre": r.get("NOMBRE"), "motivo": m}
            for r, m in zip(dup_rows, dup_motivos)
        ],
        "errors": [
            {"usuario": r.get("USUARIO"), "nombre": r.get("NOMBRE"), "motivo": m}
            for r, m in zip(error_rows, error_motivos)
        ],
        "report_token": report_token,
    }


def sync_usuarios_from_excel(file_obj, db) -> dict:
    df, err = _validate_and_parse_usuarios(file_obj)
    if err:
        return {"error": err, "updated": 0, "inserted": 0, "errors": [], "report_token": None}

    existing = {r[0]: r[1] for r in db.execute(
        "SELECT usuario, is_admin FROM usuarios"
    ).fetchall()}
    error_rows, error_motivos = [], []
    updated = inserted = 0
    user_rows: dict = {}

    for _, row in df.iterrows():
        usuario = _sv(row['USUARIO'])
        if not usuario:
            rr = {c: _sv(row.get(c)) for c in _USER_REPORT_COL_NAMES}
            error_rows.append(rr); error_motivos.append("Campo USUARIO vacío o inválido")
        else:
            user_rows[usuario] = row

    for usuario, row in user_rows.items():
        contrasena = _sv(row['CONTRASENA'])
        row_rep    = {c: _sv(row.get(c)) for c in _USER_REPORT_COL_NAMES}
        if not contrasena:
            error_rows.append(row_rep)
            error_motivos.append(f"CONTRASENA vacía para '{usuario}'")
            continue
        if usuario in existing and existing[usuario] == 1:
            error_rows.append(row_rep)
            error_motivos.append(f"'{usuario}' es Administrador — no puede modificarse por sincronización masiva")
            continue
        try:
            if usuario in existing:
                db.execute(
                    """UPDATE usuarios SET
                        nombre_completo=?, cedula=?, cargo=?, correo=?, regional=?,
                        perm_gestor_1=?, perm_gestor_2=?, perm_lider=?, perm_contralor=?,
                        superior_inmediato=?, activo=?
                    WHERE usuario=?""",
                    (_sv(row['NOMBRE']), _sv(row['CEDULA']),
                     _sv(row['CARGO']) or "Otro", _sv(row['CORREO ELECTRONICO']),
                     _sv(row['REGIONAL']),
                     _int_perm(row['GESTOR_1']), _int_perm(row['GESTOR_2']),
                     _int_perm(row['LIDER']), _int_perm(row['CONTRALOR']),
                     _sv(row['SUPERIOR_INMEDIATO']) or DEFAULT_SUPERIOR_INMEDIATO or None,
                     _int_perm(row['ACTIVO']), usuario),
                )
                updated += 1
            else:
                db.execute(
                    """INSERT INTO usuarios
                       (usuario, password_hash, nombre_completo, cedula, cargo, correo, regional,
                        perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor,
                        is_admin, superior_inmediato, activo)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)""",
                    (usuario, hash_password(contrasena),
                     _sv(row['NOMBRE']), _sv(row['CEDULA']),
                     _sv(row['CARGO']) or "Otro", _sv(row['CORREO ELECTRONICO']),
                     _sv(row['REGIONAL']),
                     _int_perm(row['GESTOR_1']), _int_perm(row['GESTOR_2']),
                     _int_perm(row['LIDER']), _int_perm(row['CONTRALOR']),
                     _sv(row['SUPERIOR_INMEDIATO']) or DEFAULT_SUPERIOR_INMEDIATO or None,
                     _int_perm(row['ACTIVO'])),
                )
                inserted += 1
        except Exception as exc:
            error_rows.append(row_rep); error_motivos.append(f"Error de BD: {exc}")

    db.commit()
    report_token = None
    if error_rows:
        buf = _build_report_excel(error_rows, error_motivos, _USER_REPORT_COL_NAMES)
        report_token = _store_report(buf.getvalue(), db)

    return {
        "updated": updated, "inserted": inserted,
        "errors": [
            {"usuario": r.get("USUARIO"), "nombre": r.get("NOMBRE"), "motivo": m}
            for r, m in zip(error_rows, error_motivos)
        ],
        "report_token": report_token,
    }


# ---------------------------------------------------------------------------
# init_db — detecta SQLite o PostgreSQL y crea el schema correspondiente
# ---------------------------------------------------------------------------
def _is_postgres() -> bool:
    return bool(DATABASE_URL and DATABASE_URL.startswith(("postgresql", "postgres")))


def _pg_connect():
    """Conexión directa psycopg2 para DDL en init_db."""
    import psycopg2
    url = DATABASE_URL
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    return psycopg2.connect(url)


def _pg_table_columns(cur, table: str) -> set:
    cur.execute(
        "SELECT column_name FROM information_schema.columns WHERE table_name = %s",
        (table,),
    )
    return {r[0] for r in cur.fetchall()}


def _pg_table_exists(cur, table: str) -> bool:
    cur.execute(
        "SELECT 1 FROM information_schema.tables WHERE table_name = %s", (table,)
    )
    return cur.fetchone() is not None


def init_db():
    if _is_postgres():
        _init_db_postgres()
    else:
        _init_db_sqlite()


def _init_db_postgres():
    """Crea / migra el schema en PostgreSQL."""
    conn = _pg_connect()
    conn.autocommit = False
    cur = conn.cursor()

    # ── campos ───────────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS campos (
            id SERIAL PRIMARY KEY,
            codigo TEXT NOT NULL UNIQUE,
            nombre TEXT NOT NULL,
            modo TEXT NOT NULL DEFAULT 'MANUAL',
            origen TEXT,
            tipo_dato TEXT DEFAULT 'Texto',
            comentario TEXT,
            formula TEXT,
            rol TEXT NOT NULL,
            orden INTEGER DEFAULT 0,
            requerido_crear INTEGER DEFAULT 0,
            requerido_g2_lider INTEGER DEFAULT 0,
            requerido_contralor INTEGER DEFAULT 0,
            dependencias TEXT
        )
    """)

    # ── seed campos PRIMERO (necesario para crear registros con tipos correctos)
    for c in CAMPOS_DEFAULT:
        cur.execute(
            """INSERT INTO campos
               (codigo, nombre, modo, origen, tipo_dato, comentario, formula,
                rol, orden, requerido_crear, requerido_g2_lider, requerido_contralor, dependencias)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
               ON CONFLICT (codigo) DO NOTHING""",
            (c["codigo"], c["nombre"], c["modo"], c.get("origen"),
             c.get("tipo_dato","Texto"), c.get("comentario"), c.get("formula"),
             c["rol"], c.get("orden",0), c.get("requerido_crear",0),
             c.get("requerido_g2_lider",0), c.get("requerido_contralor",0),
             c.get("dependencias")),
        )

    # Migración: B pasa a MANUAL, C pasa a MANUAL (ambos editables; C controla AG y A)
    cur.execute("UPDATE campos SET modo = 'MANUAL' WHERE codigo = 'B' AND modo = 'AUTOMATICA'")
    cur.execute("UPDATE campos SET modo = 'MANUAL' WHERE codigo = 'C' AND modo != 'MANUAL'")

    _refresh_globals()

    cur.execute("SELECT codigo, tipo_dato FROM campos ORDER BY orden")
    _campo_rows_list = cur.fetchall()
    _campo_codes    = [r[0] for r in _campo_rows_list] if _campo_rows_list else list(CODE_TO_COLNAME.keys())
    _campo_tipo_map = {r[0]: r[1] for r in _campo_rows_list}

    # ── registros ─────────────────────────────────────────────────────────────
    if not _pg_table_exists(cur, "registros"):
        col_defs = "\n".join(
            f'    {_col(cod)} {_campo_sql_type(_campo_tipo_map.get(cod, "Texto"))},'
            for cod in _campo_codes
        ).rstrip(",")
        cur.execute(f"""
            CREATE TABLE registros (
                id SERIAL PRIMARY KEY,
                rol TEXT NOT NULL,
                usuario TEXT NOT NULL,
                fecha_creacion TEXT NOT NULL,
{col_defs}
            )
        """)
    else:
        existing = _pg_table_columns(cur, "registros")
        for cod in _campo_codes:
            col_name  = _col(cod).lower()
            expected  = _campo_sql_type(_campo_tipo_map.get(cod, "Texto"))
            if col_name not in existing:
                try:
                    cur.execute(f"ALTER TABLE registros ADD COLUMN {col_name} {expected}")
                except Exception:
                    conn.rollback()
            elif expected != "TEXT":
                # Corregir tipo si la columna fue creada como TEXT por error
                cur.execute(
                    "SELECT data_type FROM information_schema.columns "
                    "WHERE table_name='registros' AND column_name=%s", (col_name,)
                )
                type_row = cur.fetchone()
                if type_row and type_row[0].upper() == "TEXT":
                    try:
                        cur.execute(
                            f"ALTER TABLE registros ALTER COLUMN {col_name} "
                            f"TYPE {expected} USING {col_name}::{expected}"
                        )
                    except Exception:
                        conn.rollback()

    # Migraciones: columnas de validación en registros
    existing_reg_v = _pg_table_columns(cur, "registros")
    for _vcol_name, _vcol_type in [
        ("validado",         "INTEGER DEFAULT 0"),
        ("fecha_validacion", "TEXT"),
        ("validado_por",     "TEXT"),
    ]:
        if _vcol_name not in existing_reg_v:
            try:
                cur.execute(f"ALTER TABLE registros ADD COLUMN {_vcol_name} {_vcol_type}")
            except Exception:
                conn.rollback()

    # ── usuarios ──────────────────────────────────────────────────────────────
    if not _pg_table_exists(cur, "usuarios"):
        cur.execute("""
            CREATE TABLE usuarios (
                id SERIAL PRIMARY KEY,
                usuario TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                nombre_completo TEXT NOT NULL,
                cedula TEXT,
                cargo TEXT,
                correo TEXT,
                regional TEXT,
                perm_gestor_1 INTEGER DEFAULT 0,
                perm_gestor_2 INTEGER DEFAULT 0,
                perm_lider INTEGER DEFAULT 0,
                perm_coordinador INTEGER DEFAULT 0,
                perm_contralor INTEGER DEFAULT 0,
                is_admin INTEGER DEFAULT 0,
                superior_inmediato TEXT,
                activo INTEGER DEFAULT 1
            )
        """)

    # ── tablas auxiliares ─────────────────────────────────────────────────────
    for tbl_sql in [
        """CREATE TABLE IF NOT EXISTS lista_opciones (
            id SERIAL PRIMARY KEY,
            codigo_campo TEXT NOT NULL,
            nombre_campo TEXT NOT NULL,
            valor TEXT NOT NULL,
            activo INTEGER DEFAULT 1,
            fecha_creacion TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS notificaciones (
            id SERIAL PRIMARY KEY,
            usuario_destino TEXT NOT NULL,
            tipo TEXT NOT NULL,
            mensaje TEXT NOT NULL,
            registro_id INTEGER,
            leida INTEGER DEFAULT 0,
            fecha_creacion TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS solicitudes_prestador (
            id SERIAL PRIMARY KEY,
            nit TEXT NOT NULL,
            comentario TEXT,
            solicitante TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            comentario_respuesta TEXT,
            fecha_solicitud TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS solicitudes_usuario (
            id SERIAL PRIMARY KEY,
            nombre_completo TEXT NOT NULL,
            correo TEXT NOT NULL,
            regional TEXT NOT NULL,
            rol_solicitado TEXT NOT NULL DEFAULT 'GESTOR 1',
            comentario TEXT,
            solicitante TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            comentario_respuesta TEXT,
            fecha_solicitud TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS auditoria_registros (
            id SERIAL PRIMARY KEY,
            registro_id INTEGER NOT NULL,
            comentario_admin TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'activa',
            admin_usuario TEXT NOT NULL,
            destinatario_usuario TEXT NOT NULL,
            comentario_respuesta TEXT,
            fecha_creacion TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS festivos (
            id SERIAL PRIMARY KEY,
            fecha TEXT NOT NULL UNIQUE
        )""",
        """CREATE TABLE IF NOT EXISTS ciudad_codigos (
            id SERIAL PRIMARY KEY,
            ciudad TEXT NOT NULL,
            codigo TEXT NOT NULL,
            activo INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS acta_codigos (
            id SERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            prefijo TEXT NOT NULL,
            activo INTEGER DEFAULT 1,
            ciudad TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            registro_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            usuario TEXT NOT NULL,
            rol TEXT,
            es_autorizado INTEGER NOT NULL DEFAULT 0,
            campos_diff TEXT,
            consecutivo TEXT,
            nombre_prestador TEXT,
            fecha TEXT NOT NULL,
            motivo_comentario TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS prestadores (
            id SERIAL PRIMARY KEY,
            nit TEXT NOT NULL,
            codigo_compania TEXT, compania TEXT, cod_plan TEXT, descripcion_plan TEXT,
            forma_contratacion TEXT, digito_verificacion TEXT, tipo_id TEXT, tipo_persona TEXT,
            relacion_eps TEXT, nombre_sucursal TEXT, codigo_sucursal TEXT, ciudad_cod_dane TEXT,
            ciudad TEXT, departamento TEXT, regional TEXT, especialidad TEXT,
            descripcion_especialidad TEXT, estado TEXT, tipo_convenio TEXT, direccion TEXT,
            telefono TEXT, extension_1 TEXT, telefono_2 TEXT, extension_2 TEXT, correo TEXT,
            fecha_inicio_portabilidad TEXT, fecha_fin_portabilidad TEXT, cod_habilitacion TEXT,
            habilitacion_sede TEXT, fecha_inicio_habilitacion TEXT,
            fecha_vencimiento_habilitacion TEXT, numero_contrato TEXT,
            fecha_inicio_convenio TEXT, fecha_fin_convenio TEXT,
            tipo_prestador TEXT, naturaleza_ips TEXT, tipo_atencion TEXT,
            premium TEXT, glosa_sostenida TEXT, prioridad_servicio TEXT,
            creado_manual INTEGER DEFAULT 0, fecha_creacion TEXT
        )""",
    ]:
        cur.execute(tbl_sql)

    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_registro ON audit_log(registro_id)")

    # ── datos iniciales ───────────────────────────────────────────────────────

    # lista_opciones (solo si la tabla está vacía)
    cur.execute("SELECT COUNT(*) FROM lista_opciones")
    if cur.fetchone()[0] == 0:
        for l in LISTAS_DEFAULT:
            cur.execute(
                "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
                " VALUES (%s,%s,%s,%s)",
                (l["codigo_campo"], l["nombre_campo"], l["valor"], l.get("activo", 1)),
            )

    # festivos (ON CONFLICT DO NOTHING por UNIQUE en fecha)
    for f in FESTIVOS_DEFAULT:
        cur.execute(
            "INSERT INTO festivos (fecha) VALUES (%s) ON CONFLICT DO NOTHING", (f,)
        )

    # Migración: opciones nuevas de BY que pueden no existir en BDs previas
    _by_nuevas = [
        ('BY', 'ESTADO ACTA CONCILIACIÓN', 'ENVIADA A CONTROLAR MEDICO NACIONAL'),
        ('BY', 'ESTADO ACTA CONCILIACIÓN', 'DEVUELTO COMO CONTRARLO PARA REVISION'),
    ]
    for _c, _n, _v in _by_nuevas:
        cur.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT %s,%s,%s,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=%s AND valor=%s"
            ")",
            (_c, _n, _v, _c, _v),
        )

    # Migración: DU pasa de Texto libre a Lista (SI / NO)
    cur.execute(
        "UPDATE campos SET origen='LISTA', tipo_dato='Texto lista'"
        " WHERE codigo='DU' AND (origen IS NULL OR tipo_dato='Texto')"
    )
    for _v in ('SI', 'NO'):
        cur.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT %s,%s,%s,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=%s AND valor=%s"
            ")",
            ('DU', 'CASO PAGO ESPECIAL/OBS URG CHM', _v, 'DU', _v),
        )

    # Migración: grupos 3-5 de devoluciones (FI-FW)
    _dev_campos = [
        ('FI','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION3','MANUAL',None,'Texto',None,None,'CONTRALOR',2001,0,0,0,None),
        ('FJ','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD3','MANUAL',None,'Fecha',None,None,'CONTRALOR',2002,0,0,0,None),
        ('FK','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2003,0,0,0,None),
        ('FL','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2004,0,0,0,None),
        ('FM','CASO PARA MATRIZ DE RIESGO 3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2005,0,0,0,None),
        ('FN','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION4','MANUAL',None,'Texto',None,None,'CONTRALOR',2006,0,0,0,None),
        ('FO','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD4','MANUAL',None,'Fecha',None,None,'CONTRALOR',2007,0,0,0,None),
        ('FP','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2008,0,0,0,None),
        ('FQ','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2009,0,0,0,None),
        ('FR','CASO PARA MATRIZ DE RIESGO 4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2010,0,0,0,None),
        ('FS','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION5','MANUAL',None,'Texto',None,None,'CONTRALOR',2011,0,0,0,None),
        ('FT','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD5','MANUAL',None,'Fecha',None,None,'CONTRALOR',2012,0,0,0,None),
        ('FU','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2013,0,0,0,None),
        ('FV','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2014,0,0,0,None),
        ('FW','CASO PARA MATRIZ DE RIESGO 5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2015,0,0,0,None),
    ]
    for _dc in _dev_campos:
        cur.execute(
            "INSERT INTO campos (codigo,nombre,modo,origen,tipo_dato,comentario,formula,rol,orden,"
            "requerido_crear,requerido_g2_lider,requerido_contralor,dependencias)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (codigo) DO NOTHING",
            _dc,
        )
    _dd_opts = ['DEVOLUCIÓN CM POR CARGUE EN BH - CON MODIFICACIÓN ACUERDO DE PAGO','INCUMPLIMIENTO ACUERDO DE PAGO - INTERVENCIÓN SNS','INCUMPLIMIENTO ACUERDO DE PAGO - DEVOLUCIÓN POR RECLASIFICACIÓN FACTURAS/CARGUE COMPLEMENTOS POR CUENTAS MEDICAS A PESAR ESTAR BLOQUEADO EL PRESTADOR','NO GENERA INCUMPLIMIENTO DE PAGO - DEVOLUCIÓN POR RECLASIFICACIÓN FACTURAS/CARGUE COMPLEMENTOS POR CUENTAS MEDICAS A PESAR ESTAR BLOQUEADO EL PRESTADOR','DEVOLUCIÓN CM POR CARGUE EN BH - SIN MODIFICACIÓN ACUERDO DE PAGO','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO-INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO- INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','DEVOLUCIÓN CON INCUMPLIMIENTO ACUERDO DE PAGO-INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','DEVOLUCIÓN CON INCUMPLIMIENTO ACUERDO DE PAGO-INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO- NO COHERENCIA EN LOS ARCHIVOS ADJUNTOS VS EL ACTA CONCILIADA','INCUMPLIMIENTO ACUERDO DE PAGO POR ENTREGA INOPORTUNA DE ACTA/MATRIZ','INCUMPLIMIENTO ACUERDO DE PAGO POR NO CARGUE DE ARCHIVOS EN LOS SISTEMAS DE INFORMACIÓN','NO ENVÍO ACTA PROCESO CONCILIACIÓN SUSCRITO','INCLUMPLIMIENTO ACUERDO DE PAGO - TESORERIA','QUEJA DEL PROCESO DE LA CONCILIACIÓN','NO COHERENCIA EN LOS ARCHIVOS ADJUNTOS VS EL ACTA CONCILIADA','INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA CORREO DE APROBACIÓN VS EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS ADJUNTOS','LIDER DE LA REGIONAL SOLICITA REABRIR CONCILIACIÓN, LA HABIA CERRADO LA REGIONAL COMO "CERRADO SIN RESPUESTA2','DEVOLUCION ACTA CON INCUMPLIMIENTO PAGO - SE ACEPTA GLOSA USUARIO CAPITADO','RETROALIMENTACIÓN INFORMACIÓN REGISTRO RUR/PROCESO DE CONCILIACIÓN REALIZADO','INCUMPLIMIENTO DE ACUERDO DE PAGO POR INCONSISTENCIA EN LA PARAMÉTRICA DE LA IPS, EN LOS SISTEMAS DE INFORMACIÓN DE LA EPS','NO APLICA','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL RUR']
    _de_opts = ['AUXILIAR OPERATIVO','SNS - TESORERIA','CONTRALOR NACIONAL','CUENTAS MEDICAS','CUENTAS MEDICAS/RESPONSABLE CONCILIACIÓN','LIDER REGIONAL','LIDER REGIONAL/ RESPONSABLE DE LA CONCILIACION','LIDER REGIONAL/ RESPONSABLE DE LA CONCILIACION/ AUXILIAR OPERATIVO','LIDER REGIONAL/ AUXILIAR OPERATIVO','RESPONSABLE DE LA CONCILIACION /AUXILIAR OPERATIVO','EBS NOVEDAD EN PARAMETRIZACION IPS','RESPONSABLE DE LA CONCILIACION','NO APLICA','TESORERIA','CONVENIOS Y TARIFAS','SONDA']
    _df_opts = ['SI','NO','NO APLICA']
    _dev_lista = (
        [('FK','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN3',v) for v in _dd_opts] +
        [('FP','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN4',v) for v in _dd_opts] +
        [('FU','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN5',v) for v in _dd_opts] +
        [('FL','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN3',v) for v in _de_opts] +
        [('FQ','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN4',v) for v in _de_opts] +
        [('FV','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN5',v) for v in _de_opts] +
        [('FM','CASO PARA MATRIZ DE RIESGO 3',v) for v in _df_opts] +
        [('FR','CASO PARA MATRIZ DE RIESGO 4',v) for v in _df_opts] +
        [('FW','CASO PARA MATRIZ DE RIESGO 5',v) for v in _df_opts]
    )
    for _c, _n, _v in _dev_lista:
        cur.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT %s,%s,%s,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=%s AND valor=%s"
            ")",
            (_c, _n, _v, _c, _v),
        )

    # Helper: ejecutar DDL con savepoint para evitar InFailedSqlTransaction en PostgreSQL
    def _pg_ddl(sql, params=None):
        cur.execute("SAVEPOINT _sp_ddl")
        try:
            cur.execute(sql, params) if params else cur.execute(sql)
            cur.execute("RELEASE SAVEPOINT _sp_ddl")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT _sp_ddl")

    # Migración: columna motivo_comentario en audit_log (PostgreSQL)
    _pg_ddl("ALTER TABLE audit_log ADD COLUMN IF NOT EXISTS motivo_comentario TEXT")

    # Migración: columnas de reapertura por lider en registros (PostgreSQL)
    for _rc_col in ("reapertura_lider_ac", "reapertura_lider_bd", "reapertura_lider_ce"):
        _pg_ddl(f"ALTER TABLE registros ADD COLUMN IF NOT EXISTS {_rc_col} INTEGER DEFAULT 0")

    # Migración: columnas aprobacion_n en registros (PostgreSQL)
    for _an_col, _an_type in [
        ("estado_aprobacion_n",               "TEXT"),
        ("comentario_solicitud_n",            "TEXT"),
        ("comentario_rechazo_n",              "TEXT"),
        ("comentario_aprobacion_contralor_n", "TEXT"),
        ("n_valor_anterior",                  "TEXT"),
        ("origen_pendiente_n",                "TEXT"),
    ]:
        _pg_ddl(f"ALTER TABLE registros ADD COLUMN IF NOT EXISTS {_an_col} {_an_type}")

    # Migración: tabla aprobacion_n_log (PostgreSQL)
    _pg_ddl("""
        CREATE TABLE IF NOT EXISTS aprobacion_n_log (
            id SERIAL PRIMARY KEY,
            registro_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            usuario TEXT NOT NULL,
            comentario TEXT,
            fecha TEXT NOT NULL
        )
    """)

    # Migración: tabla temp_reports para reportes de carga (PostgreSQL)
    _pg_ddl("""
        CREATE TABLE IF NOT EXISTS temp_reports (
            token TEXT PRIMARY KEY,
            created_at DOUBLE PRECISION NOT NULL,
            data BYTEA NOT NULL
        )
    """)

    # Migración: columna proceso_finalizado en registros (PostgreSQL)
    _pg_ddl("ALTER TABLE registros ADD COLUMN IF NOT EXISTS proceso_finalizado INTEGER DEFAULT 0")

    # Migración: tabla config_umbral_lider_contralor (PostgreSQL)
    _pg_ddl("""
        CREATE TABLE IF NOT EXISTS config_umbral_lider_contralor (
            id SERIAL PRIMARY KEY,
            campo_codigo TEXT NOT NULL,
            umbral REAL NOT NULL,
            activo INTEGER DEFAULT 1
        )
    """)

    # Migración: campo FX — GLOSA PERTINENCIA/CALIDAD- EXCLUIDA CONCILIACIÓN GE (PostgreSQL)
    cur.execute(
        "INSERT INTO campos (codigo,nombre,modo,origen,tipo_dato,comentario,formula,rol,orden,"
        "requerido_crear,requerido_g2_lider,requerido_contralor,dependencias)"
        " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (codigo) DO NOTHING",
        ('FX', 'GLOSA PERTINENCIA/CALIDAD- EXCLUIDA CONCILIACIÓN GE', 'MANUAL', None, 'Moneda',
         None, None, 'GESTOR 2,LIDER', 265, 0, 0, 0, None),
    )
    cur.execute(
        "UPDATE campos SET formula='SUMA(S8:AA8;FX8;AJ8)' WHERE codigo='AD' AND formula='SUMA(S8:AA8;AJ8)'"
    )
    _pg_ddl("ALTER TABLE registros ADD COLUMN IF NOT EXISTS FX REAL")

    # Migración: tabla login_attempts para lockout por fuerza bruta (PostgreSQL)
    _pg_ddl("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id SERIAL PRIMARY KEY,
            usuario TEXT NOT NULL,
            ip TEXT,
            timestamp DOUBLE PRECISION NOT NULL
        )
    """)
    _pg_ddl("CREATE INDEX IF NOT EXISTS idx_login_attempts_usuario ON login_attempts(usuario, timestamp)")

    # Migración: columna local_auth_enabled en usuarios (EntraID SSO)
    _pg_ddl("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS local_auth_enabled INTEGER DEFAULT 0")
    # Admin siempre puede usar login con contraseña como fallback
    cur.execute("UPDATE usuarios SET local_auth_enabled = 1 WHERE usuario = 'admin' AND local_auth_enabled = 0")

    # Migración: tabla sso_access_requests (solicitudes de acceso vía EntraID)
    _pg_ddl("""
        CREATE TABLE IF NOT EXISTS sso_access_requests (
            id SERIAL PRIMARY KEY,
            email TEXT NOT NULL,
            nombre TEXT,
            comentario TEXT,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            fecha TEXT NOT NULL
        )
    """)

    # ciudad_codigos (solo si la tabla está vacía)
    cur.execute("SELECT COUNT(*) FROM ciudad_codigos")
    if cur.fetchone()[0] == 0:
        for cc in CIUDAD_CODIGOS_DEFAULT:
            cur.execute(
                "INSERT INTO ciudad_codigos (ciudad, codigo, activo) VALUES (%s, %s, 1)",
                (cc["ciudad"], cc["codigo"]),
            )

    # usuario admin por defecto
    cur.execute("SELECT COUNT(*) FROM usuarios")
    if cur.fetchone()[0] == 0:
        import secrets as _secrets
        _admin_pass = ADMIN_INITIAL_PASSWORD or _secrets.token_urlsafe(16)
        if not ADMIN_INITIAL_PASSWORD:
            print(
                f"\n{'='*60}\n"
                f"  ATENCIÓN: Usuario administrador creado\n"
                f"  Usuario:    admin\n"
                f"  Contraseña: {_admin_pass}\n"
                f"  Cambie esta contraseña inmediatamente.\n"
                f"{'='*60}\n"
            )
        cur.execute(
            """INSERT INTO usuarios
               (usuario, password_hash, nombre_completo, cargo,
                perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor,
                is_admin, activo)
               VALUES (%s, %s, 'Administrador General', 'Otro', 0, 0, 0, 0, 1, 1)""",
            ("admin", hash_password(_admin_pass)),
        )

    conn.commit()
    cur.close()
    conn.close()
    print("[init_db] Base de datos PostgreSQL inicializada correctamente.")


def _init_db_sqlite():
    db = sqlite3.connect(DB_PATH)

    db.execute("""
        CREATE TABLE IF NOT EXISTS campos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL UNIQUE,
            nombre TEXT NOT NULL,
            modo TEXT NOT NULL DEFAULT 'MANUAL',
            origen TEXT,
            tipo_dato TEXT DEFAULT 'Texto',
            comentario TEXT,
            formula TEXT,
            rol TEXT NOT NULL,
            orden INTEGER DEFAULT 0
        )
    """)
    db.commit()

    _campos_cols = {r[1] for r in db.execute("PRAGMA table_info(campos)").fetchall()}
    for _col_def in [
        ("requerido_crear",     "INTEGER DEFAULT 0"),
        ("requerido_g2_lider",  "INTEGER DEFAULT 0"),
        ("requerido_contralor", "INTEGER DEFAULT 0"),
        ("dependencias",        "TEXT"),
    ]:
        if _col_def[0] not in _campos_cols:
            db.execute(f"ALTER TABLE campos ADD COLUMN {_col_def[0]} {_col_def[1]}")
    db.commit()

    # Refrescar globals ANTES de procesar registros
    _refresh_globals()

    _campo_rows  = db.execute("SELECT codigo, tipo_dato FROM campos ORDER BY orden").fetchall()
    _campo_codes = [r[0] for r in _campo_rows] if _campo_rows else list(CODE_TO_COLNAME.keys())
    _campo_tipo_map = {r[0]: r[1] for r in _campo_rows}

    info           = db.execute("PRAGMA table_info(registros)").fetchall()
    existing_cols  = [r[1] for r in info]
    # Columnas válidas: nombres largos de CODE_TO_COLNAME + nombres cortos fallback para códigos sin entrada
    descriptive_cols = set(CODE_TO_COLNAME.values()) | {_col(cod) for cod in _campo_codes}

    if not existing_cols:
        col_defs = "\n".join(
            f'    {_col(cod)} {_campo_sql_type(_campo_tipo_map.get(cod, "Texto"))},'
            for cod in _campo_codes
        )
        col_defs = col_defs.rstrip(",")
        db.execute(f"""
            CREATE TABLE registros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rol TEXT NOT NULL,
                usuario TEXT NOT NULL,
                fecha_creacion TEXT NOT NULL,
{col_defs}
            )
        """)
        db.commit()
    elif "datos" in existing_cols or not descriptive_cols.intersection(existing_cols):
        db.row_factory = sqlite3.Row
        _migrate_registros(db)
    else:
        protected = {'id', 'rol', 'usuario', 'fecha_creacion',
                     'validado', 'fecha_validacion', 'validado_por'}
        cols_to_remove = [c for c in existing_cols if c not in descriptive_cols and c not in protected]
        existing_col_types = {r[1]: r[2] for r in info}
        needs_type_migration = any(
            existing_col_types.get(_col(cod), "TEXT") == "TEXT"
            and _campo_sql_type(_campo_tipo_map.get(cod, "Texto")) != "TEXT"
            for cod in _campo_codes
            if _col(cod) in existing_col_types
        )
        if cols_to_remove or needs_type_migration:
            db.row_factory = sqlite3.Row
            _migrate_registros(db)
        else:
            for cod in _campo_codes:
                if _col(cod) not in existing_cols:
                    sql_type = _campo_sql_type(_campo_tipo_map.get(cod, "Texto"))
                    try:
                        db.execute(f"ALTER TABLE registros ADD COLUMN {_col(cod)} {sql_type}")
                    except Exception:
                        pass
            db.commit()

    # usuarios
    usuarios_info = db.execute("PRAGMA table_info(usuarios)").fetchall()
    usuarios_cols = {row[1] for row in usuarios_info}
    if usuarios_cols and "perm_gestor_1" not in usuarios_cols:
        db.execute("DROP TABLE IF EXISTS usuarios")
        usuarios_cols = set()
    if not usuarios_cols:
        db.execute("""
            CREATE TABLE usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                nombre_completo TEXT NOT NULL,
                cedula TEXT,
                cargo TEXT,
                correo TEXT,
                regional TEXT,
                perm_gestor_1 INTEGER DEFAULT 0,
                perm_gestor_2 INTEGER DEFAULT 0,
                perm_lider INTEGER DEFAULT 0,
                perm_coordinador INTEGER DEFAULT 0,
                perm_contralor INTEGER DEFAULT 0,
                is_admin INTEGER DEFAULT 0,
                superior_inmediato TEXT,
                activo INTEGER DEFAULT 1
            )
        """)

    for tbl_sql in [
        """CREATE TABLE IF NOT EXISTS lista_opciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_campo TEXT NOT NULL,
            nombre_campo TEXT NOT NULL,
            valor TEXT NOT NULL,
            activo INTEGER DEFAULT 1,
            fecha_creacion TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS notificaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_destino TEXT NOT NULL,
            tipo TEXT NOT NULL,
            mensaje TEXT NOT NULL,
            registro_id INTEGER,
            leida INTEGER DEFAULT 0,
            fecha_creacion TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS solicitudes_prestador (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nit TEXT NOT NULL,
            comentario TEXT,
            solicitante TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            comentario_respuesta TEXT,
            fecha_solicitud TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS solicitudes_usuario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_completo TEXT NOT NULL,
            correo TEXT NOT NULL,
            regional TEXT NOT NULL,
            rol_solicitado TEXT NOT NULL DEFAULT 'GESTOR 1',
            comentario TEXT,
            solicitante TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            comentario_respuesta TEXT,
            fecha_solicitud TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS auditoria_registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registro_id INTEGER NOT NULL,
            comentario_admin TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'activa',
            admin_usuario TEXT NOT NULL,
            destinatario_usuario TEXT NOT NULL,
            comentario_respuesta TEXT,
            fecha_creacion TEXT NOT NULL,
            fecha_respuesta TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS festivos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL UNIQUE
        )""",
        """CREATE TABLE IF NOT EXISTS ciudad_codigos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ciudad TEXT NOT NULL,
            codigo TEXT NOT NULL,
            activo INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS acta_codigos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            prefijo TEXT NOT NULL,
            activo INTEGER DEFAULT 1
        )""",
        """CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registro_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            usuario TEXT NOT NULL,
            rol TEXT,
            es_autorizado INTEGER NOT NULL DEFAULT 0,
            campos_diff TEXT,
            consecutivo TEXT,
            nombre_prestador TEXT,
            fecha TEXT NOT NULL,
            motivo_comentario TEXT
        )""",
        """CREATE TABLE IF NOT EXISTS aprobacion_n_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registro_id INTEGER NOT NULL,
            accion TEXT NOT NULL,
            usuario TEXT NOT NULL,
            comentario TEXT,
            fecha TEXT NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS temp_reports (
            token TEXT PRIMARY KEY,
            created_at REAL NOT NULL,
            data BLOB NOT NULL
        )""",
        """CREATE TABLE IF NOT EXISTS config_umbral_lider_contralor (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            campo_codigo TEXT NOT NULL,
            umbral REAL NOT NULL,
            activo INTEGER DEFAULT 1
        )""",
    ]:
        db.execute(tbl_sql)
    db.execute("CREATE INDEX IF NOT EXISTS idx_audit_registro ON audit_log(registro_id)")
    db.execute("""
        CREATE TABLE IF NOT EXISTS login_attempts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario TEXT NOT NULL,
            ip TEXT,
            timestamp REAL NOT NULL
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_login_attempts_usuario ON login_attempts(usuario, timestamp)")
    db.commit()

    # Migración: columna local_auth_enabled en usuarios (EntraID SSO)
    _usu_cols = {r[1] for r in db.execute("PRAGMA table_info(usuarios)").fetchall()}
    if "local_auth_enabled" not in _usu_cols:
        db.execute("ALTER TABLE usuarios ADD COLUMN local_auth_enabled INTEGER DEFAULT 0")
        db.execute("UPDATE usuarios SET local_auth_enabled = 1 WHERE usuario = 'admin'")
        db.commit()

    # Migración: tabla sso_access_requests (solicitudes de acceso vía EntraID)
    db.execute("""
        CREATE TABLE IF NOT EXISTS sso_access_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT NOT NULL,
            nombre TEXT,
            comentario TEXT,
            estado TEXT NOT NULL DEFAULT 'pendiente',
            fecha TEXT NOT NULL
        )
    """)
    db.commit()

    # Migraciones menores
    try:
        db.execute("ALTER TABLE acta_codigos ADD COLUMN ciudad TEXT")
        db.commit()
    except Exception:
        pass

    # Migración: columna motivo_comentario en audit_log
    _audit_cols = {r[1] for r in db.execute("PRAGMA table_info(audit_log)").fetchall()}
    if "motivo_comentario" not in _audit_cols:
        try:
            db.execute("ALTER TABLE audit_log ADD COLUMN motivo_comentario TEXT")
            db.commit()
        except Exception:
            pass

    # Migración: B pasa a MANUAL, C pasa a MANUAL (ambos editables; C controla AG y A)
    db.execute("UPDATE campos SET modo = 'MANUAL' WHERE codigo = 'B' AND modo = 'AUTOMATICA'")
    db.execute("UPDATE campos SET modo = 'MANUAL' WHERE codigo = 'C' AND modo != 'MANUAL'")
    db.commit()

    # Migración: columnas de validación en registros
    _reg_cols_v = {r[1] for r in db.execute("PRAGMA table_info(registros)").fetchall()}
    for _vcol_name, _vcol_type in [
        ("validado",              "INTEGER DEFAULT 0"),
        ("fecha_validacion",      "TEXT"),
        ("validado_por",          "TEXT"),
        ("reapertura_lider_ac",   "INTEGER DEFAULT 0"),
        ("reapertura_lider_bd",            "INTEGER DEFAULT 0"),
        ("reapertura_lider_ce",            "INTEGER DEFAULT 0"),
        ("estado_aprobacion_n",            "TEXT"),
        ("comentario_solicitud_n",         "TEXT"),
        ("comentario_rechazo_n",           "TEXT"),
        ("comentario_aprobacion_contralor_n", "TEXT"),
        ("n_valor_anterior",               "TEXT"),
        ("origen_pendiente_n",             "TEXT"),
        ("proceso_finalizado",             "INTEGER DEFAULT 0"),
    ]:
        if _vcol_name not in _reg_cols_v:
            try:
                db.execute(f"ALTER TABLE registros ADD COLUMN {_vcol_name} {_vcol_type}")
            except Exception:
                pass
    db.commit()

    # Prestadores — recrear si schema antiguo
    prest_info = db.execute("PRAGMA table_info(prestadores)").fetchall()
    prest_cols = {row[1] for row in prest_info}
    if prest_cols and "compania" not in prest_cols:
        db.execute("DROP TABLE IF EXISTS prestadores")
        prest_cols = set()
    if not prest_cols:
        db.execute("""
            CREATE TABLE prestadores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nit TEXT NOT NULL,
                codigo_compania TEXT, compania TEXT, cod_plan TEXT, descripcion_plan TEXT,
                forma_contratacion TEXT, digito_verificacion TEXT, tipo_id TEXT, tipo_persona TEXT,
                relacion_eps TEXT, nombre_sucursal TEXT, codigo_sucursal TEXT, ciudad_cod_dane TEXT,
                ciudad TEXT, departamento TEXT, regional TEXT, especialidad TEXT,
                descripcion_especialidad TEXT, estado TEXT, tipo_convenio TEXT, direccion TEXT,
                telefono TEXT, extension_1 TEXT, telefono_2 TEXT, extension_2 TEXT, correo TEXT,
                fecha_inicio_portabilidad TEXT, fecha_fin_portabilidad TEXT, cod_habilitacion TEXT,
                habilitacion_sede TEXT, fecha_inicio_habilitacion TEXT,
                fecha_vencimiento_habilitacion TEXT, numero_contrato TEXT,
                fecha_inicio_convenio TEXT, fecha_fin_convenio TEXT,
                tipo_prestador TEXT, naturaleza_ips TEXT, tipo_atencion TEXT,
                premium TEXT, glosa_sostenida TEXT, prioridad_servicio TEXT,
                creado_manual INTEGER DEFAULT 0, fecha_creacion TEXT
            )
        """)

    # Datos iniciales — campos (idempotente: INSERT OR IGNORE por UNIQUE en codigo)
    for c in CAMPOS_DEFAULT:
        db.execute(
            """INSERT OR IGNORE INTO campos
               (codigo, nombre, modo, origen, tipo_dato, comentario, formula,
                rol, orden, requerido_crear, requerido_g2_lider, requerido_contralor, dependencias)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (c["codigo"], c["nombre"], c["modo"], c.get("origen"),
             c.get("tipo_dato","Texto"), c.get("comentario"), c.get("formula"),
             c["rol"], c.get("orden",0), c.get("requerido_crear",0),
             c.get("requerido_g2_lider",0), c.get("requerido_contralor",0),
             c.get("dependencias")),
        )
    db.commit()

    # Datos iniciales — lista_opciones (solo si vacía)
    if db.execute("SELECT COUNT(*) FROM lista_opciones").fetchone()[0] == 0:
        db.executemany(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " VALUES (?,?,?,?)",
            [(l["codigo_campo"], l["nombre_campo"], l["valor"], l.get("activo",1))
             for l in LISTAS_DEFAULT],
        )
        db.commit()

    # Datos iniciales — festivos
    db.executemany(
        "INSERT OR IGNORE INTO festivos (fecha) VALUES (?)",
        [(f,) for f in FESTIVOS_DEFAULT],
    )
    db.commit()

    # Migración: opciones nuevas de BY que pueden no existir en BDs previas
    _by_nuevas = [
        ('BY', 'ESTADO ACTA CONCILIACIÓN', 'ENVIADA A CONTROLAR MEDICO NACIONAL'),
        ('BY', 'ESTADO ACTA CONCILIACIÓN', 'DEVUELTO COMO CONTRARLO PARA REVISION'),
    ]
    for _c, _n, _v in _by_nuevas:
        db.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT ?,?,?,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=? AND valor=?"
            ")",
            (_c, _n, _v, _c, _v),
        )

    # Migración: DU pasa de Texto libre a Lista (SI / NO)
    db.execute(
        "UPDATE campos SET origen='LISTA', tipo_dato='Texto lista'"
        " WHERE codigo='DU' AND (origen IS NULL OR tipo_dato='Texto')"
    )
    for _v in ('SI', 'NO'):
        db.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT ?,?,?,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=? AND valor=?"
            ")",
            ('DU', 'CASO PAGO ESPECIAL/OBS URG CHM', _v, 'DU', _v),
        )

    # Migración: grupos 3-5 de devoluciones (FI-FW)
    _dev_campos = [
        ('FI','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION3','MANUAL',None,'Texto',None,None,'CONTRALOR',2001,0,0,0,None),
        ('FJ','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD3','MANUAL',None,'Fecha',None,None,'CONTRALOR',2002,0,0,0,None),
        ('FK','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2003,0,0,0,None),
        ('FL','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2004,0,0,0,None),
        ('FM','CASO PARA MATRIZ DE RIESGO 3','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2005,0,0,0,None),
        ('FN','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION4','MANUAL',None,'Texto',None,None,'CONTRALOR',2006,0,0,0,None),
        ('FO','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD4','MANUAL',None,'Fecha',None,None,'CONTRALOR',2007,0,0,0,None),
        ('FP','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2008,0,0,0,None),
        ('FQ','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2009,0,0,0,None),
        ('FR','CASO PARA MATRIZ DE RIESGO 4','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2010,0,0,0,None),
        ('FS','DEVOLUCIÓN PROCESO POR INCONSISTENCIA/NOVEDAD PROCESO CONCILIACION5','MANUAL',None,'Texto',None,None,'CONTRALOR',2011,0,0,0,None),
        ('FT','FECHA DEVOLUCIÓN/FECHA RETROALIMENTACION NOVEDAD5','MANUAL',None,'Fecha',None,None,'CONTRALOR',2012,0,0,0,None),
        ('FU','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2013,0,0,0,None),
        ('FV','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2014,0,0,0,None),
        ('FW','CASO PARA MATRIZ DE RIESGO 5','MANUAL','LISTA','Texto lista',None,None,'CONTRALOR',2015,0,0,0,None),
    ]
    for _dc in _dev_campos:
        db.execute(
            "INSERT OR IGNORE INTO campos (codigo,nombre,modo,origen,tipo_dato,comentario,formula,rol,orden,"
            "requerido_crear,requerido_g2_lider,requerido_contralor,dependencias)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            _dc,
        )
    _dd_opts = ['DEVOLUCIÓN CM POR CARGUE EN BH - CON MODIFICACIÓN ACUERDO DE PAGO','INCUMPLIMIENTO ACUERDO DE PAGO - INTERVENCIÓN SNS','INCUMPLIMIENTO ACUERDO DE PAGO - DEVOLUCIÓN POR RECLASIFICACIÓN FACTURAS/CARGUE COMPLEMENTOS POR CUENTAS MEDICAS A PESAR ESTAR BLOQUEADO EL PRESTADOR','NO GENERA INCUMPLIMIENTO DE PAGO - DEVOLUCIÓN POR RECLASIFICACIÓN FACTURAS/CARGUE COMPLEMENTOS POR CUENTAS MEDICAS A PESAR ESTAR BLOQUEADO EL PRESTADOR','DEVOLUCIÓN CM POR CARGUE EN BH - SIN MODIFICACIÓN ACUERDO DE PAGO','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO-INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO- INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','DEVOLUCIÓN CON INCUMPLIMIENTO ACUERDO DE PAGO-INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','DEVOLUCIÓN CON INCUMPLIMIENTO ACUERDO DE PAGO-INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','DEVOLUCIÓN SIN MODIFICACIÓN ACUERDO PAGO- NO COHERENCIA EN LOS ARCHIVOS ADJUNTOS VS EL ACTA CONCILIADA','INCUMPLIMIENTO ACUERDO DE PAGO POR ENTREGA INOPORTUNA DE ACTA/MATRIZ','INCUMPLIMIENTO ACUERDO DE PAGO POR NO CARGUE DE ARCHIVOS EN LOS SISTEMAS DE INFORMACIÓN','NO ENVÍO ACTA PROCESO CONCILIACIÓN SUSCRITO','INCLUMPLIMIENTO ACUERDO DE PAGO - TESORERIA','QUEJA DEL PROCESO DE LA CONCILIACIÓN','NO COHERENCIA EN LOS ARCHIVOS ADJUNTOS VS EL ACTA CONCILIADA','INCONSISTENCIA EN EL PROCESO DE LA CONCILIACIÓN','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS.','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA CORREO DE APROBACIÓN VS EN EL ACTA DE CONCILIACIÓN Y/O SUS ANEXOS ADJUNTOS','LIDER DE LA REGIONAL SOLICITA REABRIR CONCILIACIÓN, LA HABIA CERRADO LA REGIONAL COMO "CERRADO SIN RESPUESTA2','DEVOLUCION ACTA CON INCUMPLIMIENTO PAGO - SE ACEPTA GLOSA USUARIO CAPITADO','RETROALIMENTACIÓN INFORMACIÓN REGISTRO RUR/PROCESO DE CONCILIACIÓN REALIZADO','INCUMPLIMIENTO DE ACUERDO DE PAGO POR INCONSISTENCIA EN LA PARAMÉTRICA DE LA IPS, EN LOS SISTEMAS DE INFORMACIÓN DE LA EPS','NO APLICA','INCONSISTENCIA EN LA INFORMACIÓN REGISTRADA EN EL RUR']
    _de_opts = ['AUXILIAR OPERATIVO','SNS - TESORERIA','CONTRALOR NACIONAL','CUENTAS MEDICAS','CUENTAS MEDICAS/RESPONSABLE CONCILIACIÓN','LIDER REGIONAL','LIDER REGIONAL/ RESPONSABLE DE LA CONCILIACION','LIDER REGIONAL/ RESPONSABLE DE LA CONCILIACION/ AUXILIAR OPERATIVO','LIDER REGIONAL/ AUXILIAR OPERATIVO','RESPONSABLE DE LA CONCILIACION /AUXILIAR OPERATIVO','EBS NOVEDAD EN PARAMETRIZACION IPS','RESPONSABLE DE LA CONCILIACION','NO APLICA','TESORERIA','CONVENIOS Y TARIFAS','SONDA']
    _df_opts = ['SI','NO','NO APLICA']
    _dev_lista = (
        [('FK','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN3',v) for v in _dd_opts] +
        [('FP','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN4',v) for v in _dd_opts] +
        [('FU','TIPO DEVOLUCIÓN/RETROALIMENTACIÓN5',v) for v in _dd_opts] +
        [('FL','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN3',v) for v in _de_opts] +
        [('FQ','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN4',v) for v in _de_opts] +
        [('FV','RESPONSABLE DEVOLUCIÓN/RETROALIMENTACIÓN5',v) for v in _de_opts] +
        [('FM','CASO PARA MATRIZ DE RIESGO 3',v) for v in _df_opts] +
        [('FR','CASO PARA MATRIZ DE RIESGO 4',v) for v in _df_opts] +
        [('FW','CASO PARA MATRIZ DE RIESGO 5',v) for v in _df_opts]
    )
    for _c, _n, _v in _dev_lista:
        db.execute(
            "INSERT INTO lista_opciones (codigo_campo, nombre_campo, valor, activo)"
            " SELECT ?,?,?,1 WHERE NOT EXISTS ("
            "  SELECT 1 FROM lista_opciones WHERE codigo_campo=? AND valor=?"
            ")",
            (_c, _n, _v, _c, _v),
        )
    db.commit()

    # Agregar columnas FI-FW a registros si no existen (se insertan en campos después de la migración principal)
    _reg_cols_now = {r[1] for r in db.execute("PRAGMA table_info(registros)").fetchall()}
    for _cod_new, _tipo_new in [
        ('FI','Texto'), ('FJ','Fecha'), ('FK','Texto'), ('FL','Texto'), ('FM','Texto'),
        ('FN','Texto'), ('FO','Fecha'), ('FP','Texto'), ('FQ','Texto'), ('FR','Texto'),
        ('FS','Texto'), ('FT','Fecha'), ('FU','Texto'), ('FV','Texto'), ('FW','Texto'),
    ]:
        if _cod_new not in _reg_cols_now:
            try:
                db.execute(f"ALTER TABLE registros ADD COLUMN {_cod_new} {_campo_sql_type(_tipo_new)}")
            except Exception:
                pass
    db.commit()

    # Migración: campo FX — GLOSA PERTINENCIA/CALIDAD- EXCLUIDA CONCILIACIÓN GE
    db.execute(
        "INSERT OR IGNORE INTO campos (codigo,nombre,modo,origen,tipo_dato,comentario,formula,rol,orden,"
        "requerido_crear,requerido_g2_lider,requerido_contralor,dependencias)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        ('FX', 'GLOSA PERTINENCIA/CALIDAD- EXCLUIDA CONCILIACIÓN GE', 'MANUAL', None, 'Moneda',
         None, None, 'GESTOR 2,LIDER', 265, 0, 0, 0, None),
    )
    # Actualizar fórmula de AD para incluir FX en la suma
    db.execute(
        "UPDATE campos SET formula='SUMA(S8:AA8;FX8;AJ8)' WHERE codigo='AD' AND formula='SUMA(S8:AA8;AJ8)'"
    )
    _reg_cols_fx = {r[1] for r in db.execute("PRAGMA table_info(registros)").fetchall()}
    if 'FX' not in _reg_cols_fx:
        try:
            db.execute("ALTER TABLE registros ADD COLUMN FX REAL")
        except Exception:
            pass
    db.commit()

    # Datos iniciales — ciudad_codigos (solo si vacía)
    if db.execute("SELECT COUNT(*) FROM ciudad_codigos").fetchone()[0] == 0:
        db.executemany(
            "INSERT INTO ciudad_codigos (ciudad, codigo, activo) VALUES (?, ?, 1)",
            [(cc["ciudad"], cc["codigo"]) for cc in CIUDAD_CODIGOS_DEFAULT],
        )
        db.commit()

    # Datos iniciales — usuario admin
    if db.execute("SELECT COUNT(*) FROM usuarios").fetchone()[0] == 0:
        import secrets as _secrets
        _admin_pass = ADMIN_INITIAL_PASSWORD or _secrets.token_urlsafe(16)
        if not ADMIN_INITIAL_PASSWORD:
            print(
                f"\n{'='*60}\n"
                f"  ATENCIÓN: Usuario administrador creado\n"
                f"  Usuario:    admin\n"
                f"  Contraseña: {_admin_pass}\n"
                f"  Cambie esta contraseña inmediatamente.\n"
                f"{'='*60}\n"
            )
        db.execute(
            """INSERT INTO usuarios
               (usuario, password_hash, nombre_completo, cargo,
                perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor,
                is_admin, activo)
               VALUES (?, ?, 'Administrador General', 'Otro', 0, 0, 0, 0, 1, 1)""",
            ("admin", hash_password(_admin_pass)),
        )
        db.commit()

    db.commit()
    db.close()
    print("[init_db] Base de datos SQLite inicializada correctamente.")
