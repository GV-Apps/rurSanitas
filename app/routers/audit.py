"""routers/audit.py — Log de auditoría (solo ADMIN)."""
import io
import json
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill

from app.auth import require_admin
from app.database import get_db

router = APIRouter()


@router.get("/api/admin/audit")
def get_audit_log(
    consecutivo: str = Query(""),
    usuario: str = Query(""),
    accion: str = Query(""),
    tipo: str = Query(""),
    page: int = Query(1, ge=1),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    per_page = 50
    where_clauses, params = [], []

    if consecutivo.strip():
        where_clauses.append("LOWER(consecutivo) LIKE ?")
        params.append(f"%{consecutivo.strip().lower()}%")
    if usuario.strip():
        where_clauses.append("LOWER(usuario) LIKE ?")
        params.append(f"%{usuario.strip().lower()}%")
    if accion.strip():
        where_clauses.append("accion = ?")
        params.append(accion.strip())
    if tipo == "autorizado":
        where_clauses.append("es_autorizado = 1")
    elif tipo == "tercero":
        where_clauses.append("es_autorizado = 0")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    total  = db.execute(f"SELECT COUNT(*) FROM audit_log {where_sql}", params).fetchone()[0]
    offset = (page - 1) * per_page
    rows   = db.execute(
        f"SELECT * FROM audit_log {where_sql} ORDER BY fecha DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [dict(r) for r in rows],
    }


@router.get("/api/admin/audit/export")
def export_audit_log(
    consecutivo: str = Query(""),
    usuario: str = Query(""),
    accion: str = Query(""),
    tipo: str = Query(""),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    where_clauses, params = [], []
    if consecutivo.strip():
        where_clauses.append("LOWER(consecutivo) LIKE ?")
        params.append(f"%{consecutivo.strip().lower()}%")
    if usuario.strip():
        where_clauses.append("LOWER(usuario) LIKE ?")
        params.append(f"%{usuario.strip().lower()}%")
    if accion.strip():
        where_clauses.append("accion = ?")
        params.append(accion.strip())
    if tipo == "autorizado":
        where_clauses.append("es_autorizado = 1")
    elif tipo == "tercero":
        where_clauses.append("es_autorizado = 0")

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""
    rows      = db.execute(
        f"SELECT * FROM audit_log {where_sql} ORDER BY fecha DESC", params
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    headers     = ["Fecha", "ID Registro", "Consecutivo", "Prestador",
                   "Usuario", "Rol", "Tipo", "Acción", "Motivo Cierre", "Campo", "Valor Anterior", "Valor Nuevo"]
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font = header_font
        cell.fill = header_fill

    fill_auth  = PatternFill("solid", fgColor="DCFCE7")
    fill_third = PatternFill("solid", fgColor="FEF3C7")
    fill_del   = PatternFill("solid", fgColor="FEE2E2")

    row_num = 2
    for r in rows:
        tipo_label   = "Autorizado" if r["es_autorizado"] else "Tercero"
        accion_label = r["accion"].capitalize()
        row_fill     = (fill_del if r["accion"] == "eliminacion"
                        else fill_auth if r["es_autorizado"] else fill_third)
        try:
            diff = json.loads(r["campos_diff"] or "{}")
        except Exception:
            diff = {}

        motivo_val = r["motivo_comentario"] or ""
        if diff:
            for campo, vals in diff.items():
                ws.append([
                    r["fecha"], r["registro_id"], r["consecutivo"], r["nombre_prestador"],
                    r["usuario"], r["rol"], tipo_label, accion_label, motivo_val,
                    campo,
                    vals.get("antes", "") if vals.get("antes") is not None else "",
                    vals.get("despues", "") if vals.get("despues") is not None else "",
                ])
                for col_i in range(1, 13):
                    ws.cell(row=row_num, column=col_i).fill = row_fill
                row_num += 1
        else:
            ws.append([r["fecha"], r["registro_id"], r["consecutivo"], r["nombre_prestador"],
                       r["usuario"], r["rol"], tipo_label, accion_label, motivo_val, "", "", ""])
            for col_i in range(1, 13):
                ws.cell(row=row_num, column=col_i).fill = row_fill
            row_num += 1

    col_widths = [22, 12, 20, 30, 16, 16, 12, 14, 30, 40, 30, 30]
    for col_i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"auditoria_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})
