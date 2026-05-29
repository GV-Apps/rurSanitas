"""routers/prestadores.py — Gestión de prestadores (admin + usuario)."""
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.auth import require_admin, require_login
from app.core.helpers import (
    _PREST_COL_NAMES, _build_report_excel, _reports, _store_report,
    cargar_prestadores, crear_notificacion, sync_prestadores_from_excel,
)
from app.database import get_db
from openpyxl.styles import Font, PatternFill

router = APIRouter()

_XLSX_MAGIC = b"PK\x03\x04"       # OOXML (.xlsx) es un ZIP
_XLS_MAGIC  = b"\xD0\xCF\x11\xE0" # Legacy .xls (CFBF)


def _validate_excel_magic(file) -> None:
    header = file.read(8)
    file.seek(0)
    if not (header.startswith(_XLSX_MAGIC) or header.startswith(_XLS_MAGIC)):
        raise HTTPException(
            status_code=400,
            detail="El archivo no es un Excel válido (tipo de contenido incorrecto)",
        )


# ── Admin: formato / cargar / sincronizar / reporte ─────────────────────────

@router.get("/api/admin/prestadores/formato")
def api_admin_formato_prestadores(sess: dict = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prestadores"
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(_PREST_COL_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Formato_Prestadores.xlsx"})


@router.post("/api/admin/prestadores/cargar")
def api_admin_cargar_prestadores(
    archivo: UploadFile = File(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")
    _validate_excel_magic(archivo.file)
    result = cargar_prestadores(archivo.file, db)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.post("/api/admin/prestadores/sincronizar")
def api_admin_sincronizar_prestadores(
    archivo: UploadFile = File(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")
    _validate_excel_magic(archivo.file)
    result = sync_prestadores_from_excel(archivo.file, db)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.get("/api/admin/prestadores/reporte/{token}")
def api_admin_reporte_prestadores(token: str, db=Depends(get_db),
                                   sess: dict = Depends(require_admin)):
    row = db.execute("SELECT data FROM temp_reports WHERE token = ?", (token,)).fetchone()
    if row:
        db.execute("DELETE FROM temp_reports WHERE token = ?", (token,))
        db.commit()
        return StreamingResponse(io.BytesIO(bytes(row["data"])),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Informe_Prestadores.xlsx"})
    # Fallback a memoria (compatibilidad)
    entry = _reports.pop(token, None)
    if entry is None:
        raise HTTPException(status_code=404, detail="Reporte no encontrado o expirado")
    _, excel_bytes = entry
    return StreamingResponse(io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Informe_Prestadores.xlsx"})


@router.get("/api/admin/prestadores/exportar")
def api_admin_exportar_prestadores(ids: str = "", db=Depends(get_db),
                                    sess: dict = Depends(require_admin)):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if id_list:
        placeholders = ",".join("?" * len(id_list))
        rows = db.execute(
            f"SELECT id, nit, digito_verificacion, nombre_sucursal, ciudad, departamento, "
            f"regional, estado, tipo_prestador, tipo_persona, compania, creado_manual, fecha_creacion "
            f"FROM prestadores WHERE id IN ({placeholders}) ORDER BY nombre_sucursal",
            id_list,
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, nit, digito_verificacion, nombre_sucursal, ciudad, departamento, "
            "regional, estado, tipo_prestador, tipo_persona, compania, creado_manual, fecha_creacion "
            "FROM prestadores ORDER BY nombre_sucursal"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prestadores"
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = [
        "ID", "NIT", "Dígito verif.", "Nombre / Razón Social", "Ciudad", "Departamento",
        "Regional", "Estado", "Tipo prestador", "Tipo persona", "Compañía",
        "Creado manual", "Fecha creación",
    ]
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font

    bool_map = {0: "No", 1: "Sí"}
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1,  value=r["id"])
        ws.cell(row=row_idx, column=2,  value=r["nit"])
        ws.cell(row=row_idx, column=3,  value=r["digito_verificacion"])
        ws.cell(row=row_idx, column=4,  value=r["nombre_sucursal"])
        ws.cell(row=row_idx, column=5,  value=r["ciudad"])
        ws.cell(row=row_idx, column=6,  value=r["departamento"])
        ws.cell(row=row_idx, column=7,  value=r["regional"])
        ws.cell(row=row_idx, column=8,  value=r["estado"])
        ws.cell(row=row_idx, column=9,  value=r["tipo_prestador"])
        ws.cell(row=row_idx, column=10, value=r["tipo_persona"])
        ws.cell(row=row_idx, column=11, value=r["compania"])
        ws.cell(row=row_idx, column=12, value=bool_map.get(r["creado_manual"], r["creado_manual"]))
        ws.cell(row=row_idx, column=13, value=r["fecha_creacion"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Prestadores.xlsx"})


@router.get("/api/admin/prestadores")
def api_admin_listar_prestadores(
    q: str = Query(""),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    q = q.strip()
    if q:
        like = f"%{q}%"
        rows = db.execute(
            """SELECT * FROM prestadores
               WHERE nit LIKE ? OR nombre_sucursal LIKE ? OR ciudad LIKE ? OR regional LIKE ?
                  OR compania LIKE ? OR tipo_prestador LIKE ?
               ORDER BY nombre_sucursal LIMIT 200""",
            (like, like, like, like, like, like),
        ).fetchall()
    else:
        rows = db.execute("SELECT * FROM prestadores ORDER BY nombre_sucursal LIMIT 200").fetchall()
    return [dict(r) for r in rows]


@router.post("/api/admin/prestadores")
def api_admin_crear_prestador(
    body: dict = Body(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    nit    = (body.get("nit") or "").strip()
    nombre = (body.get("nombre_sucursal") or "").strip()
    if not nit or not nombre:
        raise HTTPException(status_code=400, detail="NIT y nombre son requeridos")
    if db.execute("SELECT id FROM prestadores WHERE nit = ?", (nit,)).fetchone():
        raise HTTPException(status_code=409, detail="Ya existe un prestador con ese NIT")

    db.execute(
        """INSERT INTO prestadores
           (nit, codigo_compania, compania, cod_plan, descripcion_plan, forma_contratacion,
            digito_verificacion, tipo_id, tipo_persona, relacion_eps,
            nombre_sucursal, codigo_sucursal, ciudad_cod_dane, ciudad, departamento, regional,
            especialidad, descripcion_especialidad, estado, tipo_convenio,
            direccion, telefono, extension_1, telefono_2, extension_2, correo,
            fecha_inicio_portabilidad, fecha_fin_portabilidad, cod_habilitacion,
            habilitacion_sede, fecha_inicio_habilitacion, fecha_vencimiento_habilitacion,
            numero_contrato, fecha_inicio_convenio, fecha_fin_convenio,
            tipo_prestador, naturaleza_ips, tipo_atencion, premium, glosa_sostenida,
            prioridad_servicio, creado_manual, fecha_creacion)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?)""",
        (nit,
         body.get("codigo_compania") or None, body.get("compania") or None,
         body.get("cod_plan") or None, body.get("descripcion_plan") or None,
         body.get("forma_contratacion") or None, body.get("digito_verificacion") or None,
         body.get("tipo_id") or None, body.get("tipo_persona") or None,
         body.get("relacion_eps") or None, nombre,
         body.get("codigo_sucursal") or None, body.get("ciudad_cod_dane") or None,
         body.get("ciudad") or None, body.get("departamento") or None,
         body.get("regional") or None, body.get("especialidad") or None,
         body.get("descripcion_especialidad") or None,
         body.get("estado") or "ACTIVO", body.get("tipo_convenio") or None,
         body.get("direccion") or None, body.get("telefono") or None,
         body.get("extension_1") or None, body.get("telefono_2") or None,
         body.get("extension_2") or None, body.get("correo") or None,
         body.get("fecha_inicio_portabilidad") or None, body.get("fecha_fin_portabilidad") or None,
         body.get("cod_habilitacion") or None, body.get("habilitacion_sede") or None,
         body.get("fecha_inicio_habilitacion") or None, body.get("fecha_vencimiento_habilitacion") or None,
         body.get("numero_contrato") or None, body.get("fecha_inicio_convenio") or None,
         body.get("fecha_fin_convenio") or None, body.get("tipo_prestador") or None,
         body.get("naturaleza_ips") or None, body.get("tipo_atencion") or None,
         body.get("premium") or None, body.get("glosa_sostenida") or None,
         body.get("prioridad_servicio") or None, datetime.now().isoformat()),
    )
    db.commit()
    return {"mensaje": "Prestador creado exitosamente"}


@router.delete("/api/admin/prestadores/{pid}")
def api_admin_eliminar_prestador(
    pid: int,
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    row = db.execute("SELECT id, nit FROM prestadores WHERE id = ?", (pid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Prestador no encontrado")

    from app.core import fields as F
    col_nit = F._col("E")
    en_uso = db.execute(  # nosemgrep
        f"SELECT COUNT(*) FROM registros WHERE {col_nit} = ?", (row["nit"],)
    ).fetchone()[0]
    if en_uso:
        raise HTTPException(
            status_code=409,
            detail=f"No se puede eliminar: el prestador NIT {row['nit']} está referenciado en {en_uso} registro(s).",
        )

    db.execute("DELETE FROM prestadores WHERE id = ?", (pid,))
    db.commit()
    return {"mensaje": "Prestador eliminado exitosamente"}


@router.put("/api/admin/prestadores/{pid}")
def api_admin_actualizar_prestador(
    pid: int,
    body: dict = Body(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not db.execute("SELECT id FROM prestadores WHERE id = ?", (pid,)).fetchone():
        raise HTTPException(status_code=404, detail="Prestador no encontrado")
    fields = {
        "nombre_sucursal": (body.get("nombre_sucursal") or "").strip() or None,
        "codigo_compania": body.get("codigo_compania") or None,
        "compania": body.get("compania") or None,
        "cod_plan": body.get("cod_plan") or None,
        "descripcion_plan": body.get("descripcion_plan") or None,
        "forma_contratacion": body.get("forma_contratacion") or None,
        "digito_verificacion": body.get("digito_verificacion") or None,
        "tipo_id": body.get("tipo_id") or None,
        "tipo_persona": body.get("tipo_persona") or None,
        "relacion_eps": body.get("relacion_eps") or None,
        "codigo_sucursal": body.get("codigo_sucursal") or None,
        "ciudad_cod_dane": body.get("ciudad_cod_dane") or None,
        "ciudad": body.get("ciudad") or None,
        "departamento": body.get("departamento") or None,
        "regional": body.get("regional") or None,
        "especialidad": body.get("especialidad") or None,
        "descripcion_especialidad": body.get("descripcion_especialidad") or None,
        "estado": body.get("estado") or "ACTIVO",
        "tipo_convenio": body.get("tipo_convenio") or None,
        "direccion": body.get("direccion") or None,
        "telefono": body.get("telefono") or None,
        "extension_1": body.get("extension_1") or None,
        "telefono_2": body.get("telefono_2") or None,
        "extension_2": body.get("extension_2") or None,
        "correo": body.get("correo") or None,
        "fecha_inicio_portabilidad": body.get("fecha_inicio_portabilidad") or None,
        "fecha_fin_portabilidad": body.get("fecha_fin_portabilidad") or None,
        "cod_habilitacion": body.get("cod_habilitacion") or None,
        "habilitacion_sede": body.get("habilitacion_sede") or None,
        "fecha_inicio_habilitacion": body.get("fecha_inicio_habilitacion") or None,
        "fecha_vencimiento_habilitacion": body.get("fecha_vencimiento_habilitacion") or None,
        "numero_contrato": body.get("numero_contrato") or None,
        "fecha_inicio_convenio": body.get("fecha_inicio_convenio") or None,
        "fecha_fin_convenio": body.get("fecha_fin_convenio") or None,
        "tipo_prestador": body.get("tipo_prestador") or None,
        "naturaleza_ips": body.get("naturaleza_ips") or None,
        "tipo_atencion": body.get("tipo_atencion") or None,
        "premium": body.get("premium") or None,
        "glosa_sostenida": body.get("glosa_sostenida") or None,
        "prioridad_servicio": body.get("prioridad_servicio") or None,
    }
    sets = ", ".join(f"{k} = ?" for k in fields)
    db.execute(f"UPDATE prestadores SET {sets} WHERE id = ?", (*fields.values(), pid))  # nosemgrep
    db.commit()
    return {"mensaje": "Prestador actualizado exitosamente"}


# ── Rutas de usuario (autocompletar, buscar, crear, solicitudes) ─────────────

@router.get("/api/prestadores/autocompletar/{prefijo}")
def autocompletar_prestador(prefijo: str, db=Depends(get_db),
                             sess: dict = Depends(require_login)):
    rows = db.execute(
        """SELECT DISTINCT nit, nombre_sucursal, ciudad, regional FROM prestadores
           WHERE nit LIKE ? ORDER BY nit LIMIT 15""",
        (prefijo + "%",),
    ).fetchall()
    return [{"nit": r["nit"], "nombre": r["nombre_sucursal"],
             "ciudad": r["ciudad"], "regional": r["regional"]} for r in rows]


@router.get("/api/prestadores/buscar/{nit}")
def buscar_prestador(nit: str, db=Depends(get_db), sess: dict = Depends(require_login)):
    rows = db.execute(
        "SELECT DISTINCT nit, digito_verificacion, nombre_sucursal, ciudad, departamento, "
        "regional, estado, tipo_prestador, tipo_persona FROM prestadores WHERE nit = ?",
        (nit,),
    ).fetchall()
    if not rows:
        return {"encontrado": False, "prestadores": []}
    return {
        "encontrado": True,
        "prestadores": [
            {"nit": r["nit"], "digito_verificacion": r["digito_verificacion"],
             "nombre": r["nombre_sucursal"], "ciudad": r["ciudad"],
             "departamento": r["departamento"], "regional": r["regional"],
             "estado": r["estado"], "tipo_prestador": r["tipo_prestador"],
             "tipo_persona": r["tipo_persona"]} for r in rows
        ],
    }


@router.get("/api/prestador_por_nit")
def api_prestador_por_nit(nit: str = Query(""), db=Depends(get_db),
                           sess: dict = Depends(require_login)):
    if not nit.strip():
        raise HTTPException(status_code=400, detail="NIT requerido")
    row = db.execute(
        """SELECT nit, nombre_sucursal, ciudad, ciudad_cod_dane, departamento,
                  regional, tipo_persona, tipo_prestador, tipo_atencion, naturaleza_ips,
                  premium, glosa_sostenida, prioridad_servicio
           FROM prestadores WHERE nit = ? LIMIT 1""",
        (nit.strip(),)
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Prestador no encontrado")
    return dict(row)


@router.post("/api/prestadores")
def crear_prestador(body: dict = Body(...), db=Depends(get_db),
                    sess: dict = Depends(require_login)):
    nit    = (body.get("nit") or "").strip()
    nombre = (body.get("nombre") or "").strip()
    if not nit or not nombre:
        raise HTTPException(status_code=400, detail="NIT y nombre son requeridos")
    if db.execute("SELECT nit FROM prestadores WHERE nit = ?", (nit,)).fetchone():
        raise HTTPException(status_code=409, detail="El NIT ya existe en la base de datos")
    db.execute(
        """INSERT INTO prestadores
           (nit, digito_verificacion, nombre_sucursal, ciudad, departamento,
            regional, estado, direccion, telefono, correo, tipo_prestador,
            creado_manual, fecha_creacion)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)""",
        (nit, body.get("digito_verificacion", ""), nombre,
         body.get("ciudad", ""), body.get("departamento", ""),
         body.get("regional", ""), "ACTIVO",
         body.get("direccion", ""), body.get("telefono", ""),
         body.get("correo", ""), body.get("tipo_prestador", ""),
         datetime.now().isoformat()),
    )
    db.commit()
    return {"mensaje": "Prestador creado exitosamente", "nit": nit}


@router.get("/api/gestores-por-regional/{regional}")
def gestores_por_regional(regional: str, db=Depends(get_db),
                           sess: dict = Depends(require_login)):
    rows = db.execute(
        """SELECT nombre_completo FROM usuarios
           WHERE UPPER(TRIM(regional)) = UPPER(TRIM(?)) AND activo = 1
             AND (perm_gestor_1 = 1 OR perm_gestor_2 = 1 OR perm_lider = 1)
           ORDER BY nombre_completo""",
        (regional,),
    ).fetchall()
    return [r["nombre_completo"] for r in rows]


# ── Solicitudes de prestador ─────────────────────────────────────────────────

@router.post("/api/solicitudes-prestador")
def crear_solicitud_prestador(body: dict = Body(...), db=Depends(get_db),
                               sess: dict = Depends(require_login)):
    nit        = (body.get("nit") or "").strip()
    comentario = (body.get("comentario") or "").strip()
    if not nit:
        raise HTTPException(status_code=400, detail="El NIT es requerido")

    solicitante = sess["usuario"]
    now         = datetime.now().isoformat()
    db.execute(
        "INSERT INTO solicitudes_prestador (nit, comentario, solicitante, estado, fecha_solicitud)"
        " VALUES (?, ?, ?, 'pendiente', ?)",
        (nit, comentario or None, solicitante, now),
    )
    db.commit()
    sol_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    me = db.execute(
        "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
        (solicitante,)
    ).fetchone()
    nombre_sol = me["nombre_completo"] if me else solicitante
    admins = db.execute("SELECT usuario FROM usuarios WHERE is_admin = 1 AND activo = 1").fetchall()
    msg_admin = f"{nombre_sol} solicita la creación del prestador con NIT {nit}." + (
        f" Comentario: {comentario}" if comentario else "")
    for adm in admins:
        crear_notificacion(db, adm["usuario"], "solicitud_prestador", msg_admin)
    db.commit()
    return {"mensaje": "Solicitud enviada correctamente", "id": sol_id}


@router.get("/api/solicitudes-prestador")
def listar_solicitudes_prestador(db=Depends(get_db), sess: dict = Depends(require_login)):
    if sess.get("is_admin"):
        rows = db.execute(
            "SELECT * FROM solicitudes_prestador ORDER BY fecha_solicitud DESC"
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM solicitudes_prestador WHERE solicitante = ? ORDER BY fecha_solicitud DESC",
            (sess["usuario"],),
        ).fetchall()
    result = []
    for r in rows:
        sol_info = db.execute(
            "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
            (r["solicitante"],)
        ).fetchone()
        result.append({
            "id": r["id"], "nit": r["nit"], "comentario": r["comentario"],
            "solicitante": r["solicitante"],
            "nombre_solicitante": sol_info["nombre_completo"] if sol_info else r["solicitante"],
            "estado": r["estado"], "comentario_respuesta": r["comentario_respuesta"],
            "fecha_solicitud": r["fecha_solicitud"], "fecha_respuesta": r["fecha_respuesta"],
        })
    return result


@router.put("/api/solicitudes-prestador/{sid}/responder")
def responder_solicitud_prestador(sid: int, body: dict = Body(...),
                                   db=Depends(get_db), sess: dict = Depends(require_login)):
    if not sess.get("is_admin"):
        raise HTTPException(status_code=403, detail="Solo el administrador puede responder solicitudes")

    estado          = body.get("estado")
    comentario_resp = (body.get("comentario_respuesta") or "").strip()
    if estado not in ("realizado", "denegado"):
        raise HTTPException(status_code=400, detail="Estado inválido")
    if estado == "denegado" and not comentario_resp:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo de la denegación")

    row = db.execute("SELECT * FROM solicitudes_prestador WHERE id = ?", (sid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if row["estado"] != "pendiente":
        raise HTTPException(status_code=409, detail="La solicitud ya fue respondida")

    now = datetime.now().isoformat()
    db.execute(
        "UPDATE solicitudes_prestador SET estado = ?, comentario_respuesta = ?, fecha_respuesta = ? WHERE id = ?",
        (estado, comentario_resp or None, now, sid),
    )
    db.commit()

    solicitante = row["solicitante"]
    nit         = row["nit"]
    if estado == "realizado":
        tipo_notif = "solicitud_realizada"
        msg_sol    = f"Tu solicitud de creación del prestador NIT {nit} fue aprobada."
    else:
        tipo_notif = "solicitud_denegada"
        msg_sol    = f"Tu solicitud de creación del prestador NIT {nit} fue denegada. Motivo: {comentario_resp}"
    crear_notificacion(db, solicitante, tipo_notif, msg_sol)

    sol_info = db.execute(
        "SELECT nombre_completo, superior_inmediato FROM usuarios WHERE usuario = ? AND activo = 1",
        (solicitante,)
    ).fetchone()
    if sol_info and sol_info["superior_inmediato"] and sol_info["superior_inmediato"] != sess["usuario"]:
        lider       = sol_info["superior_inmediato"]
        nombre_sol  = sol_info["nombre_completo"] or solicitante
        msg_lider   = (f"Solicitud de {nombre_sol} para prestador NIT {nit}: "
                       f"{'aprobada' if estado == 'realizado' else 'denegada'}.")
        crear_notificacion(db, lider, tipo_notif, msg_lider)
    db.commit()
    return {"mensaje": f"Solicitud marcada como {estado}"}


@router.get("/api/solicitudes-prestador/pendientes-count")
def solicitudes_pendientes_count(db=Depends(get_db), sess: dict = Depends(require_login)):
    if sess.get("is_admin"):
        count = db.execute(
            "SELECT COUNT(*) FROM solicitudes_prestador WHERE estado = 'pendiente'"
        ).fetchone()[0]
    else:
        count = db.execute(
            "SELECT COUNT(*) FROM solicitudes_prestador WHERE solicitante = ? AND estado = 'pendiente'",
            (sess["usuario"],)
        ).fetchone()[0]
    return {"count": count}


# ── Solicitudes de usuario (Gestor) ─────────────────────────────────────────

def _tiene_permiso_solicitar_gestor(sess: dict) -> bool:
    """Devuelve True si el usuario tiene permiso de LIDER o CONTRALOR."""
    permisos = sess.get("permisos", [])
    return "LIDER" in permisos or "CONTRALOR" in permisos


@router.post("/api/solicitudes-usuario")
def crear_solicitud_usuario(body: dict = Body(...), db=Depends(get_db),
                             sess: dict = Depends(require_login)):
    if not _tiene_permiso_solicitar_gestor(sess) and not sess.get("is_admin"):
        raise HTTPException(status_code=403,
                            detail="Solo líderes y contralores pueden solicitar la creación de gestores")

    nombre_completo = (body.get("nombre_completo") or "").strip()
    correo          = (body.get("correo") or "").strip()
    regional        = (body.get("regional") or "").strip()
    rol_solicitado  = (body.get("rol_solicitado") or "GESTOR 1").strip()
    comentario      = (body.get("comentario") or "").strip()

    if not nombre_completo:
        raise HTTPException(status_code=400, detail="El nombre completo es requerido")
    if not correo:
        raise HTTPException(status_code=400, detail="El correo electrónico es requerido")
    if not regional:
        raise HTTPException(status_code=400, detail="La regional es requerida")
    if rol_solicitado not in ("GESTOR 1", "GESTOR 2"):
        raise HTTPException(status_code=400, detail="Rol solicitado inválido")

    solicitante = sess["usuario"]
    now = datetime.now().isoformat()
    db.execute(
        "INSERT INTO solicitudes_usuario"
        " (nombre_completo, correo, regional, rol_solicitado, comentario, solicitante, estado, fecha_solicitud)"
        " VALUES (?, ?, ?, ?, ?, ?, 'pendiente', ?)",
        (nombre_completo, correo, regional, rol_solicitado, comentario or None, solicitante, now),
    )
    db.commit()
    sol_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

    me = db.execute(
        "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
        (solicitante,)
    ).fetchone()
    nombre_sol = me["nombre_completo"] if me else solicitante

    admins = db.execute("SELECT usuario FROM usuarios WHERE is_admin = 1 AND activo = 1").fetchall()
    msg_admin = (
        f"{nombre_sol} solicita la creación del gestor '{nombre_completo}' ({rol_solicitado})"
        f" en la regional {regional}."
        + (f" Comentario: {comentario}" if comentario else "")
    )
    for adm in admins:
        crear_notificacion(db, adm["usuario"], "solicitud_usuario", msg_admin)
    db.commit()
    return {"mensaje": "Solicitud enviada correctamente", "id": sol_id}


@router.get("/api/solicitudes-usuario")
def listar_solicitudes_usuario(db=Depends(get_db), sess: dict = Depends(require_login)):
    if sess.get("is_admin"):
        rows = db.execute(
            "SELECT * FROM solicitudes_usuario ORDER BY fecha_solicitud DESC"
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT * FROM solicitudes_usuario WHERE solicitante = ? ORDER BY fecha_solicitud DESC",
            (sess["usuario"],),
        ).fetchall()
    result = []
    for r in rows:
        sol_info = db.execute(
            "SELECT nombre_completo FROM usuarios WHERE usuario = ? AND activo = 1",
            (r["solicitante"],)
        ).fetchone()
        result.append({
            "id": r["id"],
            "nombre_completo": r["nombre_completo"],
            "correo": r["correo"],
            "regional": r["regional"],
            "rol_solicitado": r["rol_solicitado"],
            "comentario": r["comentario"],
            "solicitante": r["solicitante"],
            "nombre_solicitante": sol_info["nombre_completo"] if sol_info else r["solicitante"],
            "estado": r["estado"],
            "comentario_respuesta": r["comentario_respuesta"],
            "fecha_solicitud": r["fecha_solicitud"],
            "fecha_respuesta": r["fecha_respuesta"],
        })
    return result


@router.put("/api/solicitudes-usuario/{sid}/responder")
def responder_solicitud_usuario(sid: int, body: dict = Body(...),
                                 db=Depends(get_db), sess: dict = Depends(require_login)):
    if not sess.get("is_admin"):
        raise HTTPException(status_code=403,
                            detail="Solo el administrador puede responder solicitudes")

    estado          = body.get("estado")
    comentario_resp = (body.get("comentario_respuesta") or "").strip()
    if estado not in ("realizado", "denegado"):
        raise HTTPException(status_code=400, detail="Estado inválido")
    if estado == "denegado" and not comentario_resp:
        raise HTTPException(status_code=400, detail="Debe indicar el motivo de la denegación")

    row = db.execute("SELECT * FROM solicitudes_usuario WHERE id = ?", (sid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    if row["estado"] != "pendiente":
        raise HTTPException(status_code=409, detail="La solicitud ya fue respondida")

    now = datetime.now().isoformat()
    db.execute(
        "UPDATE solicitudes_usuario SET estado = ?, comentario_respuesta = ?, fecha_respuesta = ? WHERE id = ?",
        (estado, comentario_resp or None, now, sid),
    )
    db.commit()

    solicitante    = row["solicitante"]
    nombre_gestor  = row["nombre_completo"]
    rol_solicitado = row["rol_solicitado"]
    regional       = row["regional"]
    correo         = row["correo"]

    if estado == "realizado":
        tipo_notif = "solicitud_usuario_realizada"
        msg_sol = (
            f"Tu solicitud de creación de gestor fue aprobada.\n"
            f"Datos: {nombre_gestor} | {correo} | {rol_solicitado} | Regional: {regional}."
        )
    else:
        tipo_notif = "solicitud_usuario_denegada"
        msg_sol = (
            f"Tu solicitud de creación del gestor '{nombre_gestor}' fue denegada."
            f" Motivo: {comentario_resp}"
        )
    crear_notificacion(db, solicitante, tipo_notif, msg_sol)
    db.commit()
    return {"mensaje": f"Solicitud marcada como {estado}"}


@router.get("/api/solicitudes-usuario/pendientes-count")
def solicitudes_usuario_pendientes_count(db=Depends(get_db), sess: dict = Depends(require_login)):
    if sess.get("is_admin"):
        count = db.execute(
            "SELECT COUNT(*) FROM solicitudes_usuario WHERE estado = 'pendiente'"
        ).fetchone()[0]
    else:
        count = db.execute(
            "SELECT COUNT(*) FROM solicitudes_usuario WHERE solicitante = ? AND estado = 'pendiente'",
            (sess["usuario"],)
        ).fetchone()[0]
    return {"count": count}
