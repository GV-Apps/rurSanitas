"""routers/festivos.py — Gestión de festivos colombianos."""
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl.styles import Font, PatternFill
import pandas as pd

from app.auth import require_admin, require_login
from app.database import get_db

router = APIRouter()


@router.get("/api/admin/festivos/formato")
def api_admin_formato_festivos(sess: dict = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Festivos"
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 18
    cell = ws.cell(row=1, column=1, value="FECHA")
    cell.fill = hdr_fill
    cell.font = hdr_font
    ex_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    for row_idx, fecha in enumerate(["01/01/2027", "19/03/2027", "02/04/2027"], 2):
        ws.cell(row=row_idx, column=1, value=fecha).fill = ex_fill
    nota = ws.cell(row=1, column=2, value="← Use formato DD/MM/AAAA  o  AAAA-MM-DD")
    nota.font = Font(italic=True, color="888888")
    ws.column_dimensions["B"].width = 38
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Formato_Festivos.xlsx"})


@router.post("/api/admin/festivos/cargar")
def api_admin_cargar_festivos(
    archivo: UploadFile = File(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")
    try:
        df = pd.read_excel(archivo.file)
        df = df.where(pd.notna(df), None)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {e}")

    col_fecha = None
    for col in df.columns:
        if str(col).strip().upper() == "FECHA":
            col_fecha = col
            break
    if col_fecha is None:
        col_fecha = df.columns[0] if df.shape[1] >= 1 else None
    if col_fecha is None:
        raise HTTPException(status_code=400, detail="El archivo debe tener al menos una columna con fechas")

    existing  = {r[0] for r in db.execute("SELECT fecha FROM festivos").fetchall()}
    insertados, duplicados, errores = 0, [], []

    for _, row in df.iterrows():
        raw = row[col_fecha]
        if raw is None:
            continue
        fecha_iso = None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                fecha_iso = datetime.strptime(str(raw).strip(), fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                pass
        if not fecha_iso:
            try:
                if hasattr(raw, "strftime"):
                    fecha_iso = raw.strftime("%Y-%m-%d")
            except Exception:
                pass
        if not fecha_iso:
            errores.append(str(raw).strip())
            continue
        if fecha_iso in existing:
            duplicados.append(fecha_iso)
            continue
        db.execute("INSERT OR IGNORE INTO festivos (fecha) VALUES (?)", (fecha_iso,))
        existing.add(fecha_iso)
        insertados += 1

    db.commit()
    return {
        "insertados": insertados,
        "duplicados": duplicados,
        "errores":    errores,
        "mensaje": (
            f"{insertados} festivo(s) agregado(s)"
            + (f", {len(duplicados)} ya existían" if duplicados else "")
            + (f", {len(errores)} con formato inválido" if errores else "") + "."
        ),
    }


@router.get("/api/festivos")
def api_listar_festivos(db=Depends(get_db), sess: dict = Depends(require_login)):
    rows = db.execute("SELECT id, fecha FROM festivos ORDER BY fecha").fetchall()
    return [dict(r) for r in rows]


@router.post("/api/admin/festivos")
def api_crear_festivo(body: dict = Body(...), db=Depends(get_db),
                      sess: dict = Depends(require_admin)):
    fecha = (body.get("fecha") or "").strip()
    if not fecha:
        raise HTTPException(status_code=400, detail="La fecha es requerida")
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    if db.execute("SELECT id FROM festivos WHERE fecha = ?", (fecha,)).fetchone():
        raise HTTPException(status_code=409, detail="Ya existe un festivo con esa fecha")
    db.execute("INSERT INTO festivos (fecha) VALUES (?)", (fecha,))
    db.commit()
    row = db.execute("SELECT id, fecha FROM festivos WHERE fecha = ?", (fecha,)).fetchone()
    return JSONResponse(content=dict(row), status_code=201)


@router.put("/api/admin/festivos/{fid}")
def api_actualizar_festivo(fid: int, body: dict = Body(...), db=Depends(get_db),
                            sess: dict = Depends(require_admin)):
    fecha = (body.get("fecha") or "").strip()
    if not fecha:
        raise HTTPException(status_code=400, detail="La fecha es requerida")
    try:
        datetime.strptime(fecha, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")
    if not db.execute("SELECT id FROM festivos WHERE id = ?", (fid,)).fetchone():
        raise HTTPException(status_code=404, detail="Festivo no encontrado")
    if db.execute("SELECT id FROM festivos WHERE fecha = ? AND id != ?", (fecha, fid)).fetchone():
        raise HTTPException(status_code=409, detail="Ya existe un festivo con esa fecha")
    db.execute("UPDATE festivos SET fecha = ? WHERE id = ?", (fecha, fid))
    db.commit()
    return {"mensaje": "Festivo actualizado", "id": fid, "fecha": fecha}


@router.delete("/api/admin/festivos/{fid}")
def api_eliminar_festivo(fid: int, db=Depends(get_db), sess: dict = Depends(require_admin)):
    if not db.execute("SELECT id FROM festivos WHERE id = ?", (fid,)).fetchone():
        raise HTTPException(status_code=404, detail="Festivo no encontrado")
    db.execute("DELETE FROM festivos WHERE id = ?", (fid,))
    db.commit()
    return {"mensaje": "Festivo eliminado"}
