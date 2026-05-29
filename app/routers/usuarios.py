"""routers/usuarios.py — Gestión de usuarios (solo ADMIN)."""
import io

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill

from app.auth import require_admin
from app.core.helpers import (
    _USER_COL_NAMES, _reports, _store_report,
    cargar_usuarios, hash_password, sync_usuarios_from_excel,
)
from app.database import get_db

router = APIRouter()

_XLSX_MAGIC = b"PK\x03\x04"       # OOXML (.xlsx) es un ZIP
_XLS_MAGIC  = b"\xD0\xCF\x11\xE0" # Legacy .xls (CFBF)


def _validate_excel_magic(file) -> None:
    header = file.read(8)
    file.seek(0)
    if not (header.startswith(_XLSX_MAGIC) or header.startswith(_XLS_MAGIC)):
        raise HTTPException(
            status_code=400,
            detail="El archivo no es un Excel válido (tipo de contenido incorrecto)",
        )


def _validate_password_strength(password: str) -> str | None:
    if len(password) < 8:
        return "La contraseña debe tener al menos 8 caracteres"
    if not any(c.isdigit() for c in password):
        return "La contraseña debe contener al menos un número"
    return None


@router.get("/api/admin/usuarios/formato")
def api_admin_formato_usuarios(sess: dict = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col_idx, name in enumerate(_USER_COL_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font
    example_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    examples = [
        ["juan.rodriguez@epssanitas.com", "Juan Rodríguez López", 12345678, "Gestor",
         "juan.rodriguez@epssanitas.com", "BOGOTA", 1, 0, 0, 0,
         "maria.garcia@epssanitas.com", 1, "prueba123"],
        ["ana.martinez@epssanitas.com", "Ana Martínez Torres", 87654321, "Líder",
         "ana.martinez@epssanitas.com", "MEDELLIN", 0, 0, 1, 0,
         "carlos.lider@epssanitas.com", 1, "prueba456"],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=val).fill = example_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Formato_Usuarios.xlsx"})


@router.post("/api/admin/usuarios/cargar")
def api_admin_cargar_usuarios(
    archivo: UploadFile = File(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")
    _validate_excel_magic(archivo.file)
    result = cargar_usuarios(archivo.file, db)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.post("/api/admin/usuarios/sincronizar")
def api_admin_sincronizar_usuarios(
    archivo: UploadFile = File(...),
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    if not archivo.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx, .xls)")
    _validate_excel_magic(archivo.file)
    result = sync_usuarios_from_excel(archivo.file, db)
    if "error" in result:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@router.get("/api/admin/usuarios/reporte/{token}")
def api_admin_reporte_usuarios(token: str, db=Depends(get_db),
                                sess: dict = Depends(require_admin)):
    row = db.execute("SELECT data FROM temp_reports WHERE token = ?", (token,)).fetchone()
    if row:
        db.execute("DELETE FROM temp_reports WHERE token = ?", (token,))
        db.commit()
        return StreamingResponse(io.BytesIO(bytes(row["data"])),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Informe_Usuarios.xlsx"})
    # Fallback a memoria (compatibilidad)
    entry = _reports.pop(token, None)
    if entry is None:
        raise HTTPException(status_code=404, detail="Reporte no encontrado o expirado")
    _, excel_bytes = entry
    return StreamingResponse(io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Informe_Usuarios.xlsx"})


@router.get("/api/admin/usuarios/superiores")
def api_superiores(
    regional: str = "",
    permiso: str = "",
    db=Depends(get_db),
    sess: dict = Depends(require_admin),
):
    """
    Devuelve la lista de candidatos a superior inmediato según el permiso del usuario a crear/editar.
      permiso=gestor  → líderes activos (misma regional primero)
      permiso=lider   → contralores activos (misma regional primero)
      permiso=contralor | admin → lista vacía
    """
    if permiso in ("contralor", "admin", ""):
        return []

    if permiso == "gestor":
        rows = db.execute(
            "SELECT usuario, nombre_completo, regional FROM usuarios "
            "WHERE perm_lider = 1 AND activo = 1 ORDER BY nombre_completo"
        ).fetchall()
    elif permiso == "lider":
        rows = db.execute(
            "SELECT usuario, nombre_completo, regional FROM usuarios "
            "WHERE perm_contralor = 1 AND activo = 1 ORDER BY nombre_completo"
        ).fetchall()
    else:
        return []

    reg = (regional or "").strip().upper()
    misma = [dict(r) for r in rows if (r["regional"] or "").upper() == reg]
    otras  = [dict(r) for r in rows if (r["regional"] or "").upper() != reg]
    return misma + otras


@router.get("/api/admin/usuarios")
def api_listar_usuarios(db=Depends(get_db), sess: dict = Depends(require_admin)):
    rows = db.execute(
        "SELECT id, usuario, nombre_completo, cedula, cargo, correo, regional, "
        "perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor, "
        "is_admin, superior_inmediato, activo, local_auth_enabled FROM usuarios ORDER BY nombre_completo"
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        # Normalizar: si la columna no existe aún (antes de migración) usar 1 para admin, 0 para el resto
        try:
            d["local_auth_enabled"] = bool(d["local_auth_enabled"])
        except KeyError:
            d["local_auth_enabled"] = bool(d.get("is_admin"))
        result.append(d)
    return result


@router.post("/api/admin/usuarios")
def api_crear_usuario(body: dict = Body(...), db=Depends(get_db),
                      sess: dict = Depends(require_admin)):
    usuario  = (body.get("usuario") or "").strip()
    nombre   = (body.get("nombre_completo") or "").strip()
    password = (body.get("password") or "").strip()

    if not usuario or not nombre or not password:
        raise HTTPException(status_code=400, detail="Usuario, nombre y contraseña son requeridos")
    pwd_error = _validate_password_strength(password)
    if pwd_error:
        raise HTTPException(status_code=400, detail=pwd_error)
    if db.execute("SELECT id FROM usuarios WHERE usuario = ?", (usuario,)).fetchone():
        raise HTTPException(status_code=409, detail="El usuario ya existe")

    db.execute(
        """INSERT INTO usuarios
           (usuario, password_hash, nombre_completo, cedula, cargo, correo, regional,
            perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor,
            is_admin, superior_inmediato, activo)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (usuario, hash_password(password), nombre,
         body.get("cedula") or None, body.get("cargo") or None,
         body.get("correo") or None, body.get("regional") or None,
         1 if body.get("perm_gestor_1") else 0,
         1 if body.get("perm_gestor_2") else 0,
         1 if body.get("perm_lider") else 0,
         1 if body.get("perm_contralor") else 0,
         1 if body.get("is_admin") else 0,
         body.get("superior_inmediato") or None,
         1 if body.get("activo", True) else 0),
    )
    db.commit()
    return {"mensaje": "Usuario creado exitosamente"}


@router.get("/api/admin/usuarios/exportar")
def api_exportar_usuarios(ids: str = "", db=Depends(get_db),
                          sess: dict = Depends(require_admin)):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if id_list:
        placeholders = ",".join("?" * len(id_list))
        rows = db.execute(
            f"SELECT id, usuario, nombre_completo, cedula, cargo, correo, regional, "
            f"perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor, "
            f"is_admin, superior_inmediato, activo FROM usuarios "
            f"WHERE id IN ({placeholders}) ORDER BY nombre_completo",
            id_list,
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT id, usuario, nombre_completo, cedula, cargo, correo, regional, "
            "perm_gestor_1, perm_gestor_2, perm_lider, perm_contralor, "
            "is_admin, superior_inmediato, activo FROM usuarios ORDER BY nombre_completo"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers = [
        "ID", "Usuario", "Nombre completo", "Cédula", "Cargo", "Correo", "Regional",
        "Gestor 1", "Gestor 2", "Líder", "Contralor", "Admin", "Superior inmediato", "Activo",
    ]
    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = hdr_fill
        cell.font = hdr_font

    bool_map = {0: "No", 1: "Sí"}
    for row_idx, r in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1,  value=r["id"])
        ws.cell(row=row_idx, column=2,  value=r["usuario"])
        ws.cell(row=row_idx, column=3,  value=r["nombre_completo"])
        ws.cell(row=row_idx, column=4,  value=r["cedula"])
        ws.cell(row=row_idx, column=5,  value=r["cargo"])
        ws.cell(row=row_idx, column=6,  value=r["correo"])
        ws.cell(row=row_idx, column=7,  value=r["regional"])
        ws.cell(row=row_idx, column=8,  value=bool_map.get(r["perm_gestor_1"], r["perm_gestor_1"]))
        ws.cell(row=row_idx, column=9,  value=bool_map.get(r["perm_gestor_2"], r["perm_gestor_2"]))
        ws.cell(row=row_idx, column=10, value=bool_map.get(r["perm_lider"], r["perm_lider"]))
        ws.cell(row=row_idx, column=11, value=bool_map.get(r["perm_contralor"], r["perm_contralor"]))
        ws.cell(row=row_idx, column=12, value=bool_map.get(r["is_admin"], r["is_admin"]))
        ws.cell(row=row_idx, column=13, value=r["superior_inmediato"])
        ws.cell(row=row_idx, column=14, value=bool_map.get(r["activo"], r["activo"]))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Usuarios.xlsx"})


@router.delete("/api/admin/usuarios/{uid}")
def api_eliminar_usuario(uid: int, db=Depends(get_db),
                         sess: dict = Depends(require_admin)):
    row = db.execute("SELECT id, usuario, is_admin FROM usuarios WHERE id = ?", (uid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if row["usuario"] == sess.get("usuario"):
        raise HTTPException(status_code=409, detail="No puedes eliminar tu propio usuario.")
    if row["is_admin"] == 1:
        otros = db.execute(
            "SELECT COUNT(*) FROM usuarios WHERE is_admin=1 AND activo=1 AND id != ?", (uid,)
        ).fetchone()[0]
        if otros == 0:
            raise HTTPException(status_code=409,
                detail="No se puede eliminar el único administrador activo.")
    db.execute("DELETE FROM usuarios WHERE id = ?", (uid,))
    db.commit()
    return {"mensaje": "Usuario eliminado exitosamente"}


@router.put("/api/admin/usuarios/{uid}")
def api_actualizar_usuario(uid: int, body: dict = Body(...),
                            db=Depends(get_db), sess: dict = Depends(require_admin)):
    row = db.execute("SELECT id, is_admin FROM usuarios WHERE id = ?", (uid,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    target_is_admin = row["is_admin"]
    new_is_admin    = 1 if body.get("is_admin") else 0
    if target_is_admin == 1 and new_is_admin == 0:
        otros = db.execute(
            "SELECT COUNT(*) FROM usuarios WHERE is_admin=1 AND activo=1 AND id != ?", (uid,)
        ).fetchone()[0]
        if otros == 0:
            raise HTTPException(status_code=409,
                detail="No se puede quitar el rol Administrador: es el único administrador activo.")

    # local_auth_enabled: admin siempre tiene 1 (no se puede quitar); resto según el body
    _new_local_auth = 1 if body.get("local_auth_enabled") else 0
    if row["is_admin"] == 1:
        _new_local_auth = 1  # protección: admin siempre puede usar contraseña

    fields = {
        "nombre_completo":    (body.get("nombre_completo") or "").strip() or None,
        "cedula":             body.get("cedula") or None,
        "cargo":              body.get("cargo") or None,
        "correo":             body.get("correo") or None,
        "regional":           body.get("regional") or None,
        "perm_gestor_1":      1 if body.get("perm_gestor_1") else 0,
        "perm_gestor_2":      1 if body.get("perm_gestor_2") else 0,
        "perm_lider":         1 if body.get("perm_lider") else 0,
        "perm_contralor":     1 if body.get("perm_contralor") else 0,
        "is_admin":           1 if body.get("is_admin") else 0,
        "superior_inmediato": body.get("superior_inmediato") or None,
        "activo":             1 if body.get("activo", True) else 0,
        "local_auth_enabled": _new_local_auth,
    }
    new_pass = (body.get("password") or "").strip()
    if new_pass:
        pwd_error = _validate_password_strength(new_pass)
        if pwd_error:
            raise HTTPException(status_code=400, detail=pwd_error)
        fields["password_hash"] = hash_password(new_pass)

    sets = ", ".join(f"{k} = ?" for k in fields)
    db.execute(f"UPDATE usuarios SET {sets} WHERE id = ?", (*fields.values(), uid))  # nosemgrep
    db.commit()
    return {"mensaje": "Usuario actualizado exitosamente"}
