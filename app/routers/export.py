"""routers/export.py — Exportación de registros a Excel."""
import io
from datetime import datetime

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill

from app.auth import require_login
from app.core.fields import ROLE_LOWER_ROLES, _col
from app.core.helpers import get_visibility_filter
from app.database import get_db

router = APIRouter()


@router.get("/api/registros/exportar/{rol}")
def exportar_registros(
    rol: str,
    ids: str = Query(""),
    db=Depends(get_db),
    sess: dict = Depends(require_login),
):
    # ── 1. Roles visibles para el usuario ────────────────────────────────────
    user_rol = sess.get("rol", "GESTOR 1")
    if sess.get("is_admin"):
        visible_roles = {"GESTOR 1", "GESTOR 2", "LIDER", "CONTRALOR"}
    else:
        visible_roles = set(ROLE_LOWER_ROLES.get(user_rol, [])) | {user_rol}

    # ── 2. Campos visibles, ordenados por 'orden' ─────────────────────────────
    campos_all = db.execute("SELECT codigo, nombre, rol FROM campos ORDER BY orden").fetchall()

    def _campo_visible(campo_rol_str: str) -> bool:
        return any(r.strip() in visible_roles for r in (campo_rol_str or "").split(","))

    fields_ordered = [
        (r["codigo"], r["nombre"].strip())
        for r in campos_all
        if _campo_visible(r["rol"])
    ]

    # ── 3. Reubicar cuotas 9-12 inmediatamente después de cuotas 1-8 ─────────
    G2_ANCHOR = "BX"
    CR_ANCHOR = "CR"
    G2_9_12   = {"DX", "DY", "DZ", "EA", "EB", "EC", "ED", "EE"}
    CR_9_12   = {"EF", "EG", "EH", "EI"}

    g2_extra = [(c, n) for c, n in fields_ordered if c in G2_9_12]
    cr_extra = [(c, n) for c, n in fields_ordered if c in CR_9_12]
    base     = [(c, n) for c, n in fields_ordered if c not in G2_9_12 and c not in CR_9_12]

    fields_final: list = []
    g2_inserted = cr_inserted = False
    for f in base:
        fields_final.append(f)
        if f[0] == G2_ANCHOR and g2_extra:
            fields_final.extend(g2_extra)
            g2_inserted = True
        if f[0] == CR_ANCHOR and cr_extra:
            fields_final.extend(cr_extra)
            cr_inserted = True
    if not g2_inserted:
        fields_final.extend(g2_extra)
    if not cr_inserted:
        fields_final.extend(cr_extra)

    # ── 4. Registros filtrados ────────────────────────────────────────────────
    ids_param = ids.strip()
    if ids_param:
        id_list = [int(x) for x in ids_param.split(",") if x.strip().isdigit()]
        if not id_list:
            raise HTTPException(status_code=404, detail="No hay registros para exportar")
        placeholders = ",".join("?" * len(id_list))
        rows = db.execute(
            f"SELECT * FROM registros WHERE id IN ({placeholders}) ORDER BY id DESC",
            id_list,
        ).fetchall()
    else:
        where, params = get_visibility_filter(db, sess)
        if where is None:
            rows = db.execute("SELECT * FROM registros ORDER BY id DESC").fetchall()
        else:
            rows = db.execute(
                f"SELECT * FROM registros WHERE {where} ORDER BY id DESC", params
            ).fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="No hay registros para exportar")

    # ── 5. Construir Excel ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")

    meta_headers = ["ID", "Consecutivo", "Rol Creador", "Usuario", "Fecha Creación"]
    data_headers = [nombre for _, nombre in fields_final]
    all_headers  = meta_headers + data_headers

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    meta_cod_consecutivo = _col("A")
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        try:
            ws.cell(row=row_idx, column=2, value=row[meta_cod_consecutivo])
        except Exception:
            ws.cell(row=row_idx, column=2, value="")
        ws.cell(row=row_idx, column=3, value=row["rol"])
        ws.cell(row=row_idx, column=4, value=row["usuario"])
        ws.cell(row=row_idx, column=5, value=row["fecha_creacion"])

        for col_offset, (cod, _) in enumerate(fields_final, start=6):
            try:
                val = row[_col(cod)]
            except (IndexError, KeyError):
                val = None
            ws.cell(row=row_idx, column=col_offset, value=val if val is not None else "")

    # Ancho automático de columnas (máx 60)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"registros_{rol.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
