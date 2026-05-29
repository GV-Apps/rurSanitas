"""
test_prestadores.py — Tests CRUD de prestadores (admin).
Cubre: listar, crear, actualizar, eliminar, autocompletar, buscar, formato,
cargar, sincronizar, reporte.
"""
import io
import time
import os
import pytest
import openpyxl

from app.core.helpers import _PREST_COL_NAMES, _reports

ADMIN_PASS = os.environ.get("ADMIN_INITIAL_PASSWORD", "Admin1234Test!")

_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_NIT_BASE = "800100200"
_PREST_BASE = {
    "nit": _NIT_BASE,
    "nombre_sucursal": "CLINICA TEST SA",
    "ciudad": "BOGOTA",
    "regional": "BOGOTA",
    "estado": "ACTIVO",
}


@pytest.fixture(scope="module")
def admin(client):
    client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})
    yield client


@pytest.fixture(scope="module")
def prestador_id(admin):
    """Crea un prestador y retorna su id."""
    resp = admin.post("/api/admin/prestadores", json=_PREST_BASE)
    assert resp.status_code in (200, 409), f"No se pudo crear prestador: {resp.text}"
    if resp.status_code == 409:
        rows = admin.get("/api/admin/prestadores?q=CLINICA+TEST+SA").json()
        return rows[0]["id"]
    return None  # delete tests create their own


# ── Listar ────────────────────────────────────────────────────────────────────

class TestListarPrestadores:
    def test_listar_admin_ok(self, admin):
        resp = admin.get("/api/admin/prestadores")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_listar_con_busqueda(self, admin):
        resp = admin.get("/api/admin/prestadores?q=CLINICA")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_listar_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.get("/api/admin/prestadores")
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})


# ── Crear ─────────────────────────────────────────────────────────────────────

class TestCrearPrestador:
    def test_crear_ok(self, ac):
        resp = ac.post(
            "/api/admin/prestadores",
            json={"nit": "900001111", "nombre_sucursal": "NUEVO PRESTADOR TEST"},
        )
        assert resp.status_code == 200
        assert "exitosamente" in resp.json()["mensaje"].lower()

    def test_crear_sin_nit(self, ac):
        resp = ac.post(
            "/api/admin/prestadores",
            json={"nombre_sucursal": "SIN NIT"},
        )
        assert resp.status_code == 400

    def test_crear_sin_nombre(self, ac):
        resp = ac.post(
            "/api/admin/prestadores",
            json={"nit": "900009999"},
        )
        assert resp.status_code == 400

    def test_crear_nit_duplicado(self, ac):
        nit = "800200300"
        ac.post("/api/admin/prestadores", json={"nit": nit, "nombre_sucursal": "PREST A"})
        resp = ac.post("/api/admin/prestadores", json={"nit": nit, "nombre_sucursal": "PREST B"})
        assert resp.status_code == 409

    def test_crear_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.post(
            "/api/admin/prestadores",
            json={"nit": "900008888", "nombre_sucursal": "SIN AUTH"},
        )
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})


# ── Actualizar ────────────────────────────────────────────────────────────────

class TestActualizarPrestador:
    def _crear_prestador(self, ac, nit, nombre):
        r = ac.post("/api/admin/prestadores", json={"nit": nit, "nombre_sucursal": nombre})
        assert r.status_code == 200
        rows = ac.get(f"/api/admin/prestadores?q={nit}").json()
        return rows[0]["id"]

    def test_actualizar_ok(self, ac):
        pid = self._crear_prestador(ac, "900005555", "PREST UPDT")
        resp = ac.put(
            f"/api/admin/prestadores/{pid}",
            json={"nombre_sucursal": "PREST UPDT NUEVO", "estado": "ACTIVO"},
        )
        assert resp.status_code == 200
        assert "actualizado" in resp.json()["mensaje"].lower()

    def test_actualizar_no_existe(self, ac):
        resp = ac.put(
            "/api/admin/prestadores/9999999",
            json={"nombre_sucursal": "NO EXISTE"},
        )
        assert resp.status_code == 404

    def test_actualizar_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.put("/api/admin/prestadores/1", json={"nombre_sucursal": "X"})
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})


# ── Eliminar ──────────────────────────────────────────────────────────────────

class TestEliminarPrestador:
    def _crear_prestador(self, ac, nit, nombre):
        r = ac.post("/api/admin/prestadores", json={"nit": nit, "nombre_sucursal": nombre})
        assert r.status_code == 200
        rows = ac.get(f"/api/admin/prestadores?q={nit}").json()
        return rows[0]["id"]

    def test_eliminar_ok(self, ac):
        pid = self._crear_prestador(ac, "700001111", "PREST DEL TEST")
        resp = ac.delete(f"/api/admin/prestadores/{pid}")
        assert resp.status_code == 200
        assert "eliminado" in resp.json()["mensaje"].lower()

    def test_eliminar_no_existe(self, ac):
        resp = ac.delete("/api/admin/prestadores/9999999")
        assert resp.status_code == 404

    def test_eliminar_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.delete("/api/admin/prestadores/1")
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})


# ── Autocompletar / búsqueda usuario ─────────────────────────────────────────

class TestAutocompletarPrestador:
    def test_autocompletar_ok(self, ac):
        ac.post("/api/admin/prestadores", json={"nit": "123456789", "nombre_sucursal": "AUTOCOMPLETE TEST"})
        resp = ac.get("/api/prestadores/autocompletar/123")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_autocompletar_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.get("/api/prestadores/autocompletar/123")
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})

    def test_buscar_prestadores_usuario(self, ac):
        resp = ac.get("/api/prestadores/buscar/123456789")
        assert resp.status_code == 200
        data = resp.json()
        assert "encontrado" in data

    def test_buscar_prestadores_vacio(self, ac):
        resp = ac.get("/api/prestadores/buscar/000000000")
        assert resp.status_code == 200
        assert resp.json()["encontrado"] is False


# ── Formato Excel ─────────────────────────────────────────────────────────────

class TestFormatoPrestadores:
    def test_formato_ok(self, ac):
        resp = ac.get("/api/admin/prestadores/formato")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_exportar_todos(self, ac):
        resp = ac.get("/api/admin/prestadores/exportar")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_exportar_sin_sesion(self, client):
        client.post("/api/logout")
        resp = client.get("/api/admin/prestadores/formato")
        assert resp.status_code in (401, 403)
        client.post("/api/login", json={"usuario": "admin", "password": ADMIN_PASS})


def _make_prest_xlsx(rows=None) -> bytes:
    """Crea un xlsx válido con las columnas de prestadores y filas opcionales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_PREST_COL_NAMES)
    for row in (rows or []):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _prest_row(nit: str, nombre: str = "PREST TEST") -> list:
    """Fila mínima válida para el formato de prestadores (41 columnas)."""
    row = [None] * len(_PREST_COL_NAMES)
    row[0] = nit       # NUM_ID
    row[10] = nombre   # NOMBRE_SUCURSAL
    row[18] = "ACTIVO" # ESTADO
    return row


class TestCargarPrestadores:
    def test_extension_invalida_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.txt", b"not excel", "text/plain")},
        )
        assert resp.status_code == 400

    def test_magic_bytes_invalidos_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.xlsx", b"INVALID_MAGIC", _XLSX_CT)},
        )
        assert resp.status_code == 400

    def test_columnas_incorrectas_retorna_400(self, ac):
        wb = openpyxl.Workbook()
        wb.active.append(["COL_A", "COL_B"])
        buf = io.BytesIO()
        wb.save(buf)
        resp = ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.xlsx", buf.getvalue(), _XLSX_CT)},
        )
        assert resp.status_code == 400

    def test_cargar_prestador_nuevo(self, ac):
        xlsx = _make_prest_xlsx(rows=[_prest_row("600100001", "CARGAR PREST NEW 1")])
        resp = ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "inserted" in data

    def test_cargar_prestador_duplicado_reporta_duplicados(self, ac):
        ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.xlsx", _make_prest_xlsx(rows=[_prest_row("600100002", "CARGAR DUP")]), _XLSX_CT)},
        )
        resp = ac.post(
            "/api/admin/prestadores/cargar",
            files={"archivo": ("prest.xlsx", _make_prest_xlsx(rows=[_prest_row("600100002", "CARGAR DUP 2")]), _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("duplicates", [])) > 0 or len(data.get("errors", [])) > 0



class TestSincronizarPrestadores:
    def test_extension_invalida_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/prestadores/sincronizar",
            files={"archivo": ("sync.txt", b"not excel", "text/plain")},
        )
        assert resp.status_code == 400

    def test_sincronizar_nuevo_prestador(self, ac):
        xlsx = _make_prest_xlsx(rows=[_prest_row("700200001", "SYNC PREST NEW 1")])
        resp = ac.post(
            "/api/admin/prestadores/sincronizar",
            files={"archivo": ("prest.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "inserted" in data or "updated" in data

    def test_sincronizar_actualiza_prestador_existente(self, ac):
        nit = "700200001"
        xlsx = _make_prest_xlsx(rows=[_prest_row(nit, "SYNC PREST UPDATED")])
        resp = ac.post(
            "/api/admin/prestadores/sincronizar",
            files={"archivo": ("prest.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data.get("updated", 0) > 0



class TestReportePrestadores:
    def test_token_inexistente_retorna_404(self, ac):
        resp = ac.get("/api/admin/prestadores/reporte/token-inexistente-prest-123")
        assert resp.status_code == 404

    def test_token_en_memoria_retorna_excel(self, ac):
        """Reporte disponible vía _reports en memoria (fallback sin DB)."""
        token = "test_token_rpt_prest_" + str(int(time.time()))
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        _reports[token] = (time.time(), buf.getvalue())
        resp = ac.get(f"/api/admin/prestadores/reporte/{token}")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")
