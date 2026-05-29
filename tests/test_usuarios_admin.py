"""
test_usuarios_admin.py — Tests para los endpoints admin de usuarios no cubiertos.
Cubre: formato, cargar, sincronizar, reporte, superiores, exportar, delete, put.
"""
import io
import time

import openpyxl
import pytest

from app.core.helpers import _USER_COL_NAMES, _reports


def _make_usuarios_xlsx(rows=None) -> bytes:
    """Crea un xlsx válido con las columnas de usuarios y filas opcionales."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_USER_COL_NAMES)
    for row in (rows or []):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestFormato:
    def test_formato_descarga_excel(self, ac):
        resp = ac.get("/api/admin/usuarios/formato")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_formato_requiere_admin(self, client):
        client.post("/api/logout")
        resp = client.get("/api/admin/usuarios/formato")
        assert resp.status_code in (401, 403)


class TestSuperiores:
    def test_permiso_vacio_retorna_lista_vacia(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_contralor_retorna_lista_vacia(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=contralor")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_admin_retorna_lista_vacia(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=admin")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_permiso_invalido_retorna_lista_vacia(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=invalido")
        assert resp.status_code == 200
        assert resp.json() == []

    def test_gestor_retorna_lista(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=gestor")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_lider_retorna_lista(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=lider")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)

    def test_gestor_con_regional_retorna_lista(self, ac):
        resp = ac.get("/api/admin/usuarios/superiores?permiso=gestor&regional=BOGOTA")
        assert resp.status_code == 200
        assert isinstance(resp.json(), list)


class TestExportar:
    def test_exportar_todos(self, ac):
        resp = ac.get("/api/admin/usuarios/exportar")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_exportar_por_id_valido(self, ac):
        usuarios = ac.get("/api/admin/usuarios").json()
        admin_id = next(u["id"] for u in usuarios if u["usuario"] == "admin")
        resp = ac.get(f"/api/admin/usuarios/exportar?ids={admin_id}")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")

    def test_exportar_ids_no_numericos_exporta_todos(self, ac):
        resp = ac.get("/api/admin/usuarios/exportar?ids=abc,xyz")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")


class TestDeleteUsuario:
    def test_delete_usuario_ok(self, ac):
        ac.post("/api/admin/usuarios", json={
            "usuario": "del_target_1",
            "nombre_completo": "Delete Target 1",
            "password": "DelPass1234",
        })
        uid = next(
            (u["id"] for u in ac.get("/api/admin/usuarios").json()
             if u["usuario"] == "del_target_1"),
            None,
        )
        assert uid is not None
        resp = ac.delete(f"/api/admin/usuarios/{uid}")
        assert resp.status_code == 200

    def test_delete_not_found(self, ac):
        resp = ac.delete("/api/admin/usuarios/99999999")
        assert resp.status_code == 404

    def test_delete_self_retorna_409(self, ac):
        admin_id = next(
            u["id"] for u in ac.get("/api/admin/usuarios").json()
            if u["usuario"] == "admin"
        )
        resp = ac.delete(f"/api/admin/usuarios/{admin_id}")
        assert resp.status_code == 409

    def test_delete_otro_admin_con_admin_restante(self, ac):
        """Borrar un admin secundario es válido cuando el admin principal sigue activo."""
        ac.post("/api/admin/usuarios", json={
            "usuario": "admin_sec_del",
            "nombre_completo": "Admin Secundario Del",
            "password": "SecAdmin1234",
            "is_admin": True,
        })
        uid = next(
            (u["id"] for u in ac.get("/api/admin/usuarios").json()
             if u["usuario"] == "admin_sec_del"),
            None,
        )
        assert uid is not None
        resp = ac.delete(f"/api/admin/usuarios/{uid}")
        assert resp.status_code == 200


class TestUpdateUsuario:
    def test_update_not_found(self, ac):
        resp = ac.put("/api/admin/usuarios/99999999", json={"nombre_completo": "X"})
        assert resp.status_code == 404

    def test_update_password_debil_retorna_400(self, ac):
        uid = ac.get("/api/admin/usuarios").json()[0]["id"]
        resp = ac.put(f"/api/admin/usuarios/{uid}", json={
            "nombre_completo": "Test",
            "is_admin": True,
            "password": "corta",
        })
        assert resp.status_code == 400

    def test_update_usuario_ok(self, ac):
        ac.post("/api/admin/usuarios", json={
            "usuario": "upd_target_1",
            "nombre_completo": "Update Target 1",
            "password": "UpdPass1234",
        })
        uid = next(
            (u["id"] for u in ac.get("/api/admin/usuarios").json()
             if u["usuario"] == "upd_target_1"),
            None,
        )
        assert uid is not None
        resp = ac.put(f"/api/admin/usuarios/{uid}", json={
            "nombre_completo": "Updated Name",
            "regional": "MEDELLIN",
            "perm_gestor_1": True,
            "activo": True,
            "is_admin": False,
            "local_auth_enabled": True,
        })
        assert resp.status_code == 200

    def test_update_con_nueva_password(self, ac):
        ac.post("/api/admin/usuarios", json={
            "usuario": "upd_pwd_1",
            "nombre_completo": "Pwd Update",
            "password": "InitPass1234",
        })
        uid = next(
            (u["id"] for u in ac.get("/api/admin/usuarios").json()
             if u["usuario"] == "upd_pwd_1"),
            None,
        )
        assert uid is not None
        resp = ac.put(f"/api/admin/usuarios/{uid}", json={
            "nombre_completo": "Pwd Update",
            "password": "NewPass5678",
        })
        assert resp.status_code == 200

    def test_update_ultimo_admin_no_puede_demoter(self, ac):
        """No se puede quitar el rol admin al único administrador activo."""
        todos = ac.get("/api/admin/usuarios").json()
        for u in todos:
            if u["usuario"] != "admin" and u["is_admin"]:
                ac.delete(f"/api/admin/usuarios/{u['id']}")
        admin_id = next(
            u["id"] for u in ac.get("/api/admin/usuarios").json()
            if u["usuario"] == "admin"
        )
        resp = ac.put(f"/api/admin/usuarios/{admin_id}", json={
            "nombre_completo": "Administrador General",
            "is_admin": False,
        })
        assert resp.status_code == 409

    def test_update_admin_mantiene_local_auth(self, ac):
        """Actualizar el admin siempre conserva local_auth_enabled=1."""
        admin_id = next(
            u["id"] for u in ac.get("/api/admin/usuarios").json()
            if u["usuario"] == "admin"
        )
        resp = ac.put(f"/api/admin/usuarios/{admin_id}", json={
            "nombre_completo": "Administrador General",
            "is_admin": True,
            "local_auth_enabled": False,
        })
        assert resp.status_code == 200


class TestCargarUsuarios:
    def test_extension_invalida_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.txt", b"not excel", "text/plain")},
        )
        assert resp.status_code == 400

    def test_magic_bytes_invalidos_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", b"INVALID_MAGIC_BYTES", _XLSX_CT)},
        )
        assert resp.status_code == 400

    def test_columnas_incorrectas_retorna_400(self, ac):
        wb = openpyxl.Workbook()
        wb.active.append(["COL_A", "COL_B"])
        buf = io.BytesIO()
        wb.save(buf)
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", buf.getvalue(), _XLSX_CT)},
        )
        assert resp.status_code == 400

    def test_cargar_usuario_nuevo(self, ac):
        xlsx = _make_usuarios_xlsx(rows=[
            ["carga_new1@t.com", "Carga New1", 11000001, "Gestor",
             "carga_new1@t.com", "BOGOTA", 1, 0, 0, 0, "", 1, "CargaPass1"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "inserted" in data

    def test_cargar_usuario_duplicado_genera_reporte(self, ac):
        xlsx = _make_usuarios_xlsx(rows=[
            ["carga_new1@t.com", "Carga Dup", 11000002, "Gestor",
             "carga_new1@t.com", "BOGOTA", 1, 0, 0, 0, "", 1, "CargaDup1"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("duplicates", [])) > 0 or len(data.get("errors", [])) > 0

    def test_cargar_usuario_con_campo_usuario_vacio(self, ac):
        """Filas con USUARIO vacío van a errores, no bloquean el proceso."""
        xlsx = _make_usuarios_xlsx(rows=[
            [None, "Sin Usuario", 99000001, "Gestor", "", "BOGOTA", 0, 0, 0, 0, "", 1, "Pass1234"],
            ["valido_vacio@t.com", "Válido", 99000002, "Gestor", "", "BOGOTA", 1, 0, 0, 0, "", 1, "Pass5678"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("errors", [])) > 0

    def test_cargar_usuario_admin_va_a_errores(self, ac):
        """Intentar cargar al usuario admin por carga masiva va a errores."""
        xlsx = _make_usuarios_xlsx(rows=[
            ["admin", "Admin Carga", 0, "Admin", "admin@t.com", "BOGOTA", 0, 0, 0, 0, "", 1, "AdminPass1"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("errors", [])) > 0

    def test_cargar_usuario_duplicado_en_archivo(self, ac):
        """Mismo usuario dos veces en el mismo Excel: el segundo va a duplicados."""
        xlsx = _make_usuarios_xlsx(rows=[
            ["dup_en_arch@t.com", "Dup Arch 1", 88000001, "Gestor", "", "BOGOTA", 1, 0, 0, 0, "", 1, "DupArch1"],
            ["dup_en_arch@t.com", "Dup Arch 2", 88000002, "Gestor", "", "CALI", 1, 0, 0, 0, "", 1, "DupArch2"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/cargar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("duplicates", [])) > 0


class TestSincronizarUsuarios:
    def test_extension_invalida_retorna_400(self, ac):
        resp = ac.post(
            "/api/admin/usuarios/sincronizar",
            files={"archivo": ("sync.txt", b"not excel", "text/plain")},
        )
        assert resp.status_code == 400

    def test_sincronizar_nuevo_usuario(self, ac):
        xlsx = _make_usuarios_xlsx(rows=[
            ["sync_new1@t.com", "Sync New1", 22000001, "Gestor",
             "sync_new1@t.com", "CALI", 1, 0, 0, 0, "", 1, "SyncPass1234"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/sincronizar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "inserted" in data or "updated" in data

    def test_sincronizar_actualiza_usuario_existente(self, ac):
        xlsx = _make_usuarios_xlsx(rows=[
            ["sync_new1@t.com", "Sync New1 Updated", 22000001, "Lider",
             "sync_new1@t.com", "MEDELLIN", 0, 0, 1, 0, "", 1, "SyncUpd5678"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/sincronizar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data.get("updated", 0) > 0

    def test_sincronizar_admin_va_a_errores(self, ac):
        """Sincronizar al usuario admin falla con error protegido."""
        xlsx = _make_usuarios_xlsx(rows=[
            ["admin", "Admin Sync", 0, "Admin", "admin@t.com", "BOGOTA", 0, 0, 0, 0, "", 1, "AdminSync1"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/sincronizar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("errors", [])) > 0

    def test_sincronizar_usuario_vacio_va_a_errores(self, ac):
        """Filas con USUARIO vacío van a errores en sincronización."""
        xlsx = _make_usuarios_xlsx(rows=[
            [None, "Sin Usuario Sync", 77000001, "Gestor", "", "BOGOTA", 1, 0, 0, 0, "", 1, "SyncPass99"],
        ])
        resp = ac.post(
            "/api/admin/usuarios/sincronizar",
            files={"archivo": ("usuarios.xlsx", xlsx, _XLSX_CT)},
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data.get("errors", [])) > 0


class TestReporteUsuarios:
    def test_token_inexistente_retorna_404(self, ac):
        resp = ac.get("/api/admin/usuarios/reporte/token-inexistente-xyz-123")
        assert resp.status_code == 404

    def test_token_en_memoria_retorna_excel(self, ac):
        """Reporte disponible vía _reports en memoria (fallback sin DB)."""
        token = "test_token_rpt_usuarios_" + str(int(time.time()))
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        _reports[token] = (time.time(), buf.getvalue())
        resp = ac.get(f"/api/admin/usuarios/reporte/{token}")
        assert resp.status_code == 200
        assert "spreadsheet" in resp.headers.get("content-type", "")
